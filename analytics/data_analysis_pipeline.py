import os
from datetime import datetime
import pandas as pd
from sqlalchemy import create_engine, text
import json
from contextlib import contextmanager
import re
import logging
import google.generativeai as genai
from dotenv import load_dotenv
import ast
import uuid
from typing import List, Dict, Any, Required
import io
import base64
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

load_dotenv()

log_format = "%(levelname)s: %(message)s"
logging.basicConfig(level=logging.INFO, format=log_format)
logger = logging.getLogger(__name__)


# --- Database Connection ---
@contextmanager
def get_db_connection():
    """Provides a managed database connection."""
    # Required environment variables
    db_name = os.environ.get("DB_NAME")
    db_user = os.environ.get("DB_USER")
    db_host = os.environ.get("DB_HOST")
    db_port = os.environ.get("DB_PORT")
    db_password = os.environ.get("DB_PASSWORD")
    instance_connection_name = os.environ.get("INSTANCE_CONNECTION_NAME")

    # Validate required environment variables
    if not all([db_name, db_user, db_password, instance_connection_name]):
        missing_vars = []
        if not db_name:
            missing_vars.append("DB_NAME")
        if not db_user:
            missing_vars.append("DB_USER")
        if not db_password:
            missing_vars.append("DB_PASSWORD")
        if not instance_connection_name:
            missing_vars.append("INSTANCE_CONNECTION_NAME")

        raise ValueError(
            f"Missing required environment variables: {', '.join(missing_vars)}"
        )

    from urllib.parse import quote_plus

    # For Cloud SQL Connector using Unix domain sockets
    # Format: postgresql+psycopg2://user:password@/dbname?host=/cloudsql/instance_connection_name
    # conn_string = (
    #     f"postgresql+psycopg2://{db_user}:{quote_plus(db_password)}@/{db_name}"
    #     f"?host=/cloudsql/{instance_connection_name}"
    # )

    conn_string = (
        f"postgresql+psycopg2://{db_user}:{quote_plus(db_password)}@/{db_name}"
        f"?host=35.190.189.103"
    )

    # conn_string = f"postgresql+psycopg2://{db_user}:{quote_plus(db_password)}@{db_host}:{db_port}/{db_name}"

    # Alternative format (both should work):
    # conn_string = (
    #     f"postgresql+psycopg2://{db_user}:{quote_plus(db_password)}@/{db_name}"
    #     f"?host=/cloudsql/{instance_connection_name}"
    # )

    engine = None
    conn = None
    try:
        logger.debug("Attempting to connect to the database via Cloud SQL Connector.")
        engine = create_engine(
            conn_string,
            pool_size=5,
            max_overflow=10,
            pool_timeout=30,
            pool_recycle=1800,  # 30 minutes
        )
        conn = engine.connect()
        logger.debug("Database connection successful.")
        yield conn, engine
    except Exception as e:
        logger.error(f"❌ Database connection failed: {e}", exc_info=True)
        raise
    finally:
        if conn:
            conn.close()
            logger.debug("Database connection closed.")
        if engine:
            engine.dispose()
            logger.debug("Database engine disposed.")


def sql_query_with_params(query: str, params: dict = None) -> List[Dict[str, Any]]:
    """
    Runs a SQL SELECT query on the PostgreSQL database with parameters
    and returns results as a list of dictionaries.

    Args:
        query: The SQL query string, potentially with placeholders like :param_name.
        params: A dictionary of parameters to bind to the query placeholders.

    Returns:
        A list of dictionaries representing the query results, or an empty list.

    Raises:
        Exception: If the database query fails.
    """
    logger.debug(f"Executing SQL query: {query} with params: {params}")
    try:
        with get_db_connection() as (conn, engine):
            # Use text() for the query and pass params directly to read_sql_query
            result = pd.read_sql_query(text(query), conn, params=params)
            # Convert NaN/NaT to None for JSON compatibility
            result = result.where(pd.notnull(result), None)
            data = result.to_dict(orient="records")
            logger.debug(f"Query executed successfully. Rows returned: {len(data)}")
            if data:
                # Log only a small sample if data is large
                sample_size = min(len(data), 2)
                logger.debug(
                    f"Query result sample: {json.dumps(data[:sample_size], indent=2, default=str)}"
                )
            else:
                logger.debug("Query returned no data.")
            return data
    except Exception as e:
        logger.error(f"❌ Error executing parameterized SQL query: {e}", exc_info=True)
        # Optionally return an error structure instead of raising
        # return [{"error": f"SQL execution failed: {e}"}]
        raise


def get_company_data_schema(company_id: int) -> str:
    """
    Retrieves the 'data_schema' JSONB content for a specific company
    from the 'company_data' table.

    Args:
        company_id: The identifier for the company whose schema is needed.

    Returns:
        A JSON string representing the schema stored in the 'data_schema' column,
        or an error message string starting with "Error:", or "{}" if not found/empty.
    """
    logger.debug(f"Retrieving data_schema for company_id: {company_id}")
    # Ensure company_id is treated as an integer in the query parameter
    query = text(
        "SELECT data_schema FROM company_data WHERE company_id = :company_id LIMIT 1"
    )
    schema_json = "{}"  # Default to empty JSON string

    try:
        with get_db_connection() as (conn, engine):
            # Execute query with parameter binding
            result = conn.execute(
                query, {"company_id": company_id}
            ).fetchone()

            if result and result[0]:
                # The database driver (psycopg2) usually converts JSONB to Python dict/list automatically
                schema_data = result[0]
                if (
                        isinstance(schema_data, (dict, list)) and schema_data
                ):  # Check if it's a non-empty dict/list
                    schema_json = json.dumps(schema_data, indent=2)
                    logger.debug(
                        f"Schema retrieved successfully for company_id {company_id}."
                    )
                elif isinstance(
                        schema_data, str
                ):  # Handle if it comes back as string unexpectedly
                    try:
                        parsed_schema = json.loads(schema_data)
                        if parsed_schema:
                            schema_json = json.dumps(parsed_schema, indent=2)
                            logger.debug(
                                f"Schema retrieved (parsed from string) for company_id {company_id}."
                            )
                        else:
                            logger.warning(
                                f"Empty data_schema found (after parsing string) for company_id: {company_id}"
                            )
                    except json.JSONDecodeError:
                        logger.error(
                            f"Invalid JSON string in data_schema for company_id: {company_id}"
                        )
                        return f"Error: Invalid JSON found in data_schema for company_id {company_id}"
                else:
                    logger.warning(
                        f"Empty or non-dict/list data_schema found for company_id: {company_id}"
                    )

            else:
                logger.warning(
                    f"No data_schema record found for company_id: {company_id}"
                )
                # Decide if this is an error or just means no schema available
                # Returning an error might be safer if schema is expected
                return f"Error: No data_schema found for company_id {company_id}"

    except Exception as e:
        logger.error(
            f"❌ Error retrieving data_schema for company_id {company_id}: {e}",
            exc_info=True,
        )
        return f"Error: Failed to retrieve data schema: {e}"

    # Return "{}" only if schema was explicitly empty, otherwise return the JSON string or error
    return schema_json if schema_json != "{}" else "{}"

def get_company_column_rules(company_id: int) -> str:
    """
    Retrieves the 'rules' plain TEXT content for a specific company from the 'company_data' table.
    """
    logger.debug(f"Retrieving column rules (TEXT) for company_id: {company_id}")
    query = text("SELECT rules FROM company_data WHERE company_id = :company_id LIMIT 1")
    rules_text = ""

    try:
        with get_db_connection() as (conn, engine):
            result = conn.execute(query, {"company_id": company_id}).fetchone()

            if result and result[0]:
                rules_text = str(result[0]).strip()
                if not rules_text:
                    logger.warning(f"Rules column is present but empty for company_id: {company_id}")
            else:
                logger.warning(f"No rules found for company_id: {company_id}")
                return "Error: No rules found for company_id"
    except Exception as e:
        logger.error(f"❌ Error retrieving rules for company_id {company_id}: {e}", exc_info=True)
        return f"Error: Failed to retrieve rules: {e}"

    return rules_text



def get_company_column_description(company_id: int) -> str:
    """
    Retrieves the 'column_description' JSONB content for a specific company
    from the 'company_data' table.

    Args:
        company_id: The identifier for the company whose schema is needed.

    Returns:
        A JSON string representing the schema stored in the 'description' column,
        or an error message string starting with "Error:", or "{}" if not found/empty.
    """
    logger.debug(f"Retrieving column descriptions for company_id: {company_id}")
    # Ensure company_id is treated as an integer in the query parameter
    query = text(
        "SELECT description FROM company_data WHERE company_id = :company_id LIMIT 1"
    )
    description_json = "{}"  # Default to empty JSON string

    try:
        with get_db_connection() as (conn, engine):
            # Execute query with parameter binding
            result = conn.execute(
                query, {"company_id": company_id}
            ).fetchone()

            if result and result[0]:
                # The database driver (psycopg2) usually converts JSONB to Python dict/list automatically
                description_data = result[0]
                if (
                        isinstance(description_data, (dict, list)) and description_data
                ):  # Check if it's a non-empty dict/list
                    description_json = json.dumps(description_data, indent=2)
                    logger.debug(
                        f"Column Descriptions retrieved successfully for company_id {company_id}."
                    )
                elif isinstance(
                        description_data, str
                ):  # Handle if it comes back as string unexpectedly
                    try:
                        parsed_description = json.loads(description_data)
                        if parsed_description:
                            schema_json = json.dumps(parsed_description, indent=2)
                            logger.debug(
                                f"Descriptions retrieved (parsed from string) for company_id {company_id}."
                            )
                        else:
                            logger.warning(
                                f"Empty column description found (after parsing string) for company_id: {company_id}"
                            )
                    except json.JSONDecodeError:
                        logger.error(
                            f"Invalid JSON string in description for company_id: {company_id}"
                        )
                        return f"Error: Invalid JSON found in description for company_id {company_id}"
                else:
                    logger.warning(
                        f"Empty or non-dict/list description found for company_id: {company_id}"
                    )

            else:
                logger.warning(
                    f"No description record found for company_id: {company_id}"
                )
                # Decide if this is an error or just means no description available
                # Returning an error might be safer if description is expected
                return f"Error: No description found for company_id {company_id}"

    except Exception as e:
        logger.error(
            f"❌ Error retrieving description for company_id {company_id}: {e}",
            exc_info=True,
        )
        return f"Error: Failed to retrieve data description: {e}"

    # Return "{}" only if schema was explicitly empty, otherwise return the JSON string or error
    return description_json if description_json != "{}" else "{}"

def initialize_gemini_model(model_name="gemini-1.5-flash", system_instruction=None):
    """Initializes and configures the Gemini model."""
    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        raise ValueError("GEMINI_API_KEY must be set in environment variables.")

    genai.configure(api_key=api_key)

    generation_config = {
        "temperature": 0.0,
        "top_p": 0.95,
        "top_k": 64,
        "max_output_tokens": 8192,
        "response_mime_type": "text/plain",
    }
    safety_settings = [
        {"category": "HARM_CATEGORY_HARASSMENT", "threshold": "BLOCK_MEDIUM_AND_ABOVE"},
        {
            "category": "HARM_CATEGORY_HATE_SPEECH",
            "threshold": "BLOCK_MEDIUM_AND_ABOVE",
        },
        {
            "category": "HARM_CATEGORY_SEXUALLY_EXPLICIT",
            "threshold": "BLOCK_MEDIUM_AND_ABOVE",
        },
        {
            "category": "HARM_CATEGORY_DANGEROUS_CONTENT",
            "threshold": "BLOCK_MEDIUM_AND_ABOVE",
        },
    ]

    model_kwargs = {
        "model_name": model_name,
        "safety_settings": safety_settings,
        "generation_config": generation_config,
    }
    if system_instruction:
        model_kwargs["system_instruction"] = system_instruction

    try:
        model = genai.GenerativeModel(**model_kwargs)
        logger.debug(f"Gemini model '{model_name}' initialized successfully.")
        return model
    except Exception as e:
        logger.error(
            f"❌ Failed to initialize Gemini model '{model_name}': {e}", exc_info=True
        )
        raise


def clean_response_text(text):
    """Removes markdown code blocks and trims whitespace."""
    # Remove ```sql, ```json, ```python etc. and the closing ```
    text = re.sub(r"^```[a-zA-Z]*\s*|\s*```$", "", text, flags=re.MULTILINE)
    return text.strip()


def parse_tasks_response(response_text):
    """
    Safely parse the task decomposition response into a Python list of dictionaries.
    Handles JSON null values by converting them to Python None.
    """
    try:
        # First try ast.literal_eval after replacing null/true/false
        python_compatible = (
            response_text.replace("null", "None")
            .replace("true", "True")
            .replace("false", "False")
        )
        tasks = ast.literal_eval(python_compatible)
        if not isinstance(tasks, list):
            raise ValueError("Decomposition did not return a list (evaluated by ast).")
        # Basic validation of task structure
        for task in tasks:
            if not isinstance(task, dict):
                raise ValueError(f"Task item is not a dictionary: {task}")
            if "task_type" not in task or "description" not in task:
                logger.warning(f"Task missing 'task_type' or 'description': {task}")
        return tasks
    except (SyntaxError, ValueError, TypeError) as ast_error:
        logger.warning(f"AST parsing failed: {ast_error}. Trying JSON parsing.")
        try:
            # If ast fails, try json.loads
            tasks = json.loads(response_text)
            if not isinstance(tasks, list):
                raise ValueError(
                    "Decomposition did not return a list (evaluated by json)."
                )
            # Basic validation of task structure (redundant but safe)
            for task in tasks:
                if not isinstance(task, dict):
                    raise ValueError(f"Task item is not a dictionary: {task}")
                if "task_type" not in task or "description" not in task:
                    logger.warning(f"Task missing 'task_type' or 'description': {task}")
            return tasks
        except (json.JSONDecodeError, ValueError, TypeError) as json_error:
            logger.error(
                f"Failed to parse task list. Raw response: '{response_text}'. AST error: {ast_error}. "
                f"JSON error: {json_error}",
                exc_info=False,
            )  # Keep exc_info=False for cleaner logs unless debugging heavily
            raise ValueError(
                f"Failed to parse task list. Check AI response format. Raw response snippet: '{response_text[:200]}...'"
            )


def insert_message_into_db(message_data: Dict[str, Any]) -> Dict[str, Any]:
    """Inserts a new message record into the 'messages' table."""
    query = """
        INSERT INTO messages (
            chat_id, sender, message_text, timestamp
        ) VALUES (
            :chat_id, :sender, :message_text, :timestamp
        ) RETURNING message_id;
    """

    # --- Prepare message_text as a string from a JSON object array ---
    # This is the part that converts your array of JSON objects into a string.
    # It checks if report_data or graph_data is provided and is a list/dict (JSON-like)
    # If so, it stringifies it for message_text.

    # If your LLM's 'message_text' field already contains the desired string, use that directly.
    # If not, and you want to stringify 'report_data' or 'graph_data' for 'message_text':
    _message_text = message_data.get("result", "")  # Start with what's provided

    logger.info("Inserting response into database")

    params = {
        "chat_id": message_data["chat_id"],
        "sender": message_data["sender"],
        "message_text": _message_text,  # Use the stringified version or original message_text
        "timestamp": message_data.get("timestamp", datetime.now().isoformat()),
    }

    with get_db_connection() as (conn, engine):
        result = conn.execute(text(query), params).fetchone()
        conn.commit()
        return dict(result._mapping) if result else None


def process_prompt(
        prompt: str, company_id: int, chat_id:uuid
) -> List[Dict[str, Any]]:
    """
    Processes a user prompt against data in the 'company_data' table's JSONB columns
    for a specific company.

    Args:
        prompt: The user's natural language request.
        company_id: The ID of the company whose data should be analyzed.

    Returns:
        A list of result dictionaries, each containing 'type' and 'data'.
    """
    # --- Step 1: Receive and Log Prompt & Company ID ---
    logger.info("\n✨ STEP 1: PROCESSING USER PROMPT")
    logger.info(f'Received Prompt: "{prompt}"')
    if not isinstance(company_id, int) or company_id <= 0:
        logger.error(
            f"❌ Invalid Company ID provided: {company_id}. Must be a positive integer."
        )
        return [
            {
                "type": "text",
                "data": f"Error: Invalid Company ID ({company_id}). Please provide a valid ID.",
            }
        ]
    logger.info(f"Target Company ID: {company_id}")

    results = []

    try:
        # ---- Step 2a: Get Schema ----
        logger.info("\n📜 STEP 2a: FETCHING SCHEMA FROM company_data TABLE")
        database_schema_json_or_error = get_company_data_schema(
            company_id
        )  # Use the new function

        if database_schema_json_or_error.startswith("Error:"):
            logger.error(f"❌ Schema retrieval failed: {database_schema_json_or_error}")
            return [
                {
                    "type": "text",
                    "data": f"Failed to proceed: {database_schema_json_or_error}",
                }
            ]

        if database_schema_json_or_error == "{}":
            logger.warning(
                f"Retrieved empty schema for company_id {company_id}. Cannot proceed with analysis."
            )
            return [
                {
                    "type": "text",
                    "data": f"Failed to proceed: The data schema for Company ID {company_id} is empty or could"
                            f" not be properly retrieved.",
                }
            ]

        try:
            # Validate if it's actually JSON (though get_company_data_schema should ensure this)
            schema_dict = json.loads(database_schema_json_or_error)
            if not schema_dict:  # Double check if it's empty after parsing
                logger.warning(
                    f"Database schema parsed as empty for company ID {company_id}."
                )
                return [
                    {
                        "type": "text",
                        "data": f"Failed to proceed: Parsed data schema for Company ID {company_id} is empty.",
                    }
                ]
        except json.JSONDecodeError:
            logger.error(
                f"❌ Failed to parse the retrieved schema JSON: {database_schema_json_or_error}"
            )
            return [
                {
                    "type": "text",
                    "data": "Failed to proceed: Error parsing database schema information.",
                }
            ]

        database_schema_json = (
            database_schema_json_or_error  # Use the validated JSON string
        )
        logger.info("[✓] Database schema retrieved and validated.")

        # ---- Step 2aa: Get Description ----
        logger.info("\n📜 STEP 2aa: FETCHING description FROM company_data TABLE")
        database_description_json_or_error = get_company_column_description(
            company_id
        )  # Use the new function

        if database_description_json_or_error.startswith("Error:"):
            logger.error(f"❌ Description retrieval failed: {database_description_json_or_error}")
            return [
                {
                    "type": "text",
                    "data": f"Failed to proceed: {database_description_json_or_error}",
                }
            ]

        if database_description_json_or_error == "{}":
            logger.warning(
                f"Retrieved empty description for company_id {company_id}. Cannot proceed with analysis."
            )
            return [
                {
                    "type": "text",
                    "data": f"Failed to proceed: The data description for Company ID {company_id} is empty or could"
                            f" not be properly retrieved.",
                }
            ]

        try:
            # Validate if it's actually JSON (though get_company_data_schema should ensure this)
            description_dict = json.loads(database_description_json_or_error)
            if not description_dict:  # Double check if it's empty after parsing
                logger.warning(
                    f"Database description parsed as empty for company ID {company_id}."
                )
                return [
                    {
                        "type": "text",
                        "data": f"Failed to proceed: Parsed data description for Company ID {company_id} is empty.",
                    }
                ]
        except json.JSONDecodeError:
            logger.error(
                f"❌ Failed to parse the retrieved description JSON: {database_description_json_or_error}"
            )
            return [
                {
                    "type": "text",
                    "data": "Failed to proceed: Error parsing database description information.",
                }
            ]

        database_description_json = (
            database_description_json_or_error  # Use the validated JSON string
        )
        logger.info("[✓] Database description retrieved and validated.")

        # ---- Step 2ab: Get Rules ----
        logger.info("\n📜 STEP 2ab: FETCHING rules FROM company_data TABLE")
        database_rules_text_or_error = get_company_column_rules(company_id)

        if database_rules_text_or_error.startswith("Error:"):
            logger.warning(f"[!] Rules retrieval warning: {database_rules_text_or_error}")
        elif not database_rules_text_or_error.strip():
            logger.info("Rules column is empty.")
        else:
            logger.info("[✓] Rules retrieved successfully (plain text).")

        print(database_rules_text_or_error)

        # --- Step 2b: Decompose Prompt into Tasks ---
        logger.info("\n🧠 STEP 2b: DECOMPOSING PROMPT INTO TASKS (using JSON schema)")

        # --- DECOMPOSITION INSTRUCTION (No changes needed here) ---

        decomposition_instruction = f"""
                                    Analyze the user's request: "{prompt}"
                                    Based on this request and the Data Schema provided below, identify the specific 
                                    data analysis or reporting tasks 
                                    required to fulfill the user's objectives.
                                    
                                    The data for these tasks resides in a single table 'company_data' within a JSONB 
                                    column named 'data'. You MUST filter 
                                    by company_id = {company_id}.
                                    The structure of this 'data' column for the relevant company is described by the 
                                    'Data Schema'. The keys in the 'Data 
                                    Schema' (e.g., "pms", "change_order") correspond to the top-level keys within the 
                                    'data' JSONB column, each holding 
                                    an array of JSON objects.
                                    
                                    The Description about columns of this 'data' column for the relevant company is described by the 
                                    'Description'. It's jsonb format where top key have description about it's columns
                                    
                                    Data Schema (Structure within the 'data' JSONB column of 'company_data' for 
                                    company_id={company_id}):
                                    {database_schema_json}
                                    
                                    Description about columns (Description of columns  within the 'data' JSONB column of 'company_data' for 
                                    company_id={company_id}):
                                    {database_description_json}                                    
                                    
                                    **Guidelines for Defining Tasks:**
                                    
                                    1.  **Task Identification (Understand User Objectives):**
                                        * First, carefully read the user's prompt ("{prompt}") to understand their 
                                        primary goals or the questions they 
                                        want answered.
                                        * Identify each distinct objective. An objective might be a request for 
                                        specific information, a comparison, 
                                        a summary, a trend analysis, etc.
                                        * Each distinct objective that can be addressed with the provided schema 
                                        should correspond to a task.
                                        * User can ask multiple columns within same report. consider it single report 
                                        which specifies all columns that 
                                        user want.
                                        * Consider if data from DIFFERENT KEYS within the JSONB 'data' column (e.g., 
                                        "pms" and "change_order") needs to 
                                        be conceptually combined or related to fulfill an objective.
                                    
                                    2.  **Task Type Assignment (`task_type`) and Refinement:**
                                        * For each objective identified in Step 1, assign a `task_type` as follows:
                                        * **If the user explicitly requests a specific task type** for an objective (
                                        e.g., "generate a report for X", 
                                        "create a chart of Y", "I need an insight about Z"):
                                            * The `task_type` for that objective **MUST** be what the user specified.
                                            * For such an explicitly typed request, generate **ONLY** that task for 
                                            that specific objective. Do not 
                                            generate alternative types (e.g., an insight if a report was asked for) 
                                            for the *same objective*.
                                        * **If the user does NOT explicitly state a task type** for an objective, 
                                        determine the most fitting type based 
                                        on the nature of the information sought:
                                            * **'insight':** Choose this if the objective is to understand *why* 
                                            something is happening, to identify key 
                                            takeaways, underlying patterns, significant trends (and their 
                                            implications), anomalies, or to get a 
                                            summarized interpretation or textual explanation that goes beyond just 
                                            presenting raw data. User prompts like 
                                            "What are the main drivers of sales?", "Explain the recent dip in 
                                            performance.", "Are there any unusual 
                                            activities?", "Summarize customer feedback trends.", "Why is X correlated 
                                            with Y?", "What does the data 
                                            suggest about Z?" often call for an 'insight'.
                                            * **'report':** Choose this if the objective is primarily to see 
                                            structured data, detailed listings, 
                                            specific values, or straightforward aggregations without requiring 
                                            significant interpretation or qualitative 
                                            explanation. User prompts like "List all overdue projects.", "Show me the 
                                            sales figures for each region.", 
                                            "What is the total number of users active last month?", "Generate a 
                                            summary table of X." typically call for a 
                                            'report'.
                                            * **'visualization':** Choose this if the objective is best met by a 
                                            graphical representation to see trends, 
                                            compare values, understand distributions, or highlight relationships 
                                            visually. User prompts like "Compare 
                                            sales over the last year.", "Show me the distribution of customer ages.", 
                                            "Plot X against Y." often call for 
                                            a 'visualization'. If so, also determine a `visualization_type` ('bar' or 
                                            'line').
                                            * **Decision Guidance:** When a prompt is somewhat ambiguous (e.g., 
                                            "Tell me about X," "Analyze Y"), 
                                            consider whether the core need is for raw/structured data ('report'), 
                                            a visual representation (
                                            'visualization'), or an interpretation/explanation/conclusion (
                                            'insight'). If the user seems to be asking a 
                                            "what," "why," or "so what" question that requires drawing conclusions 
                                            from data, lean towards 'insight'.
                                        * **Multiple Objectives:** If the user's prompt contains multiple distinct 
                                        objectives (e.g., "Give me a report on 
                                        A and also show a chart for B"), create a separate task for each objective, 
                                        applying these typing rules 
                                        individually.
                                    
                                    Based ONLY on the user prompt, the schema, and these guidelines, list the task(s).
                                    
                                    For each task, specify:
                                    1.  'task_type': 'insight', 'visualization', or 'report' (determined by Guideline 
                                    2).
                                    2.  'description': Brief description of the task's objective (e.g., "Report of 
                                    change orders per project manager"). 
                                    This should reflect the objective identified in Guideline 1.
                                    3.  'required_data_summary': Describe the data needed to achieve the objective, 
                                    mentioning the relevant JSON KEYS (
                                    e.g., "pms", "change_order") and the specific FIELDS from the schema (e.g., 
                                    "PM_Name from pms", "Change Orders from 
                                    change_order"). Clearly state if and how data from multiple keys needs to be 
                                    related.
                                    4.  'visualization_type': 'bar' or 'line' if `task_type` is 'visualization', 
                                    else null.
                                    
                                    Output the result as a valid Python list of dictionaries ONLY. No explanations or 
                                    markdown. Ensure keys and values 
                                    are double-quoted. Use null for missing values, not None.
                                    Example (this example shows the format; the number and nature of tasks generated 
                                    depend on the user's specific 
                                    request and the guidelines above):
                                    [
                                        {{"task_type": "report", "description": "Report linking PMs to their change 
                                        orders", "required_data_summary": 
                                        "Need PM_Name from 'pms' key and Job Number, Change Orders from 
                                        'change_order' key. Relate pms.PM_Id to 
                                        change_order.Project_Manager using extracted fields.", "visualization_type": 
                                        null}},
                                        {{"task_type": "visualization", "description": "Total change orders per PM", 
                                        "required_data_summary": "Need 
                                        PM_Name from 'pms' and Change Orders from 'change_order'. Aggregate Change 
                                        Orders grouped by PM after relating 
                                        the keys.", "visualization_type": "bar"}}
                                    ]
                                    """
        # --- End Decomposition Instruction ---

        decomposer_model = initialize_gemini_model()
        decomposer_chat = decomposer_model.start_chat()
        response = decomposer_chat.send_message(decomposition_instruction)
        cleaned_response = clean_response_text(response.text)
        logger.debug(f"Raw task decomposition response: {response.text}")
        logger.debug(f"Cleaned task decomposition response: {cleaned_response}")

        try:
            tasks = parse_tasks_response(cleaned_response)
            if not tasks:
                logger.warning(
                    "AI task decomposition returned an empty list. No tasks to perform."
                )
                return [
                    {
                        "type": "text",
                        "data": f"I couldn't identify specific tasks from your request based on the"
                                f" available data structure for Company ID {company_id}. Could you please rephrase?",
                    }
                ]

            logger.info(f"🤖 AI identified {len(tasks)} tasks:")
            for idx, task_item in enumerate(tasks):
                logger.info(
                    f"  • Task {idx + 1}: {task_item.get('description', 'N/A')} ({task_item.get('task_type', 'N/A')})"
                )
        except (ValueError, TypeError) as e:
            logger.error(
                f"❌ Failed to parse AI task decomposition: {e}", exc_info=False
            )
            return [
                {
                    "type": "text",
                    "data": f"Error: Could not understand the tasks required by"
                            f" the prompt. Please rephrase. (Parsing error: {e})",
                }
            ]

        # --- Step 3: Process Each Task ---
        logger.info("\n⚙️ STEP 3: PROCESSING TASKS")
        sql_gemini = None
        plotly_gemini = None
        insight_gemini = None
        title_gemini = None

        sql_instruction = f"""You are an expert PostgreSQL query writer specializing in JSONB data. Your task is to generate a single, syntactically correct PostgreSQL `SELECT` query based on the provided task and schema.
                                
                                **Strictly adhere to all rules below.**
                                
                                ### **Data Source**
                                * **Table:** `company_data`
                                * **Columns:**
                                    * `company_id` (INT): Unique identifier.
                                    * `data` (JSONB): Contains all uploaded file data.
                                * **Mandatory Filter:** Every CTE accessing `company_data` **must** include `WHERE company_id = :company_id`. The placeholders `:company_id` will be provided.
                                
                                ### **JSONB Schema**
                                * **Structure:** The `data` column's structure is provided as `{database_schema_json}`. It contains top-level keys (e.g., `records`, `details`) that map to arrays of JSON objects.
                                * **Fields:** Each object includes an identifier field (e.g., `id`) for joining or grouping, as specified in the task.
                                * **Example Schema:**
                                    ```json
                                    (
                                        "records": [
                                            ("id": "101", "name": "John", "value1": "1000.0", "value2": "800.0"),
                                            ("id": "102", "name": "Jane", "value1": "0.0", "value2": "0.0")
                                        ],
                                        "details": [
                                            ("id": "101", "category": "A", "amount": "50.0")
                                        ]
                                    )
                                    ```
                                * **Descriptions:** `{database_description_json}` provides column descriptions. Use this to correctly analyze the task. (May be empty.)
                                * **Business Rules:** `{database_rules_text_or_error}` specifies mandatory business rules. (May be empty.)
                                
                                ### **Querying JSONB Data**
                                
                                1.  **Unnesting Arrays:**
                                    * **Always use CTEs.** Direct unnesting in the main query or subqueries is forbidden.
                                    * **Pattern:** For each JSONB array (e.g., `data -> 'records'`), create a dedicated CTE:
                                        ```sql
                                        WITH records_data AS (
                                            SELECT company_id, jsonb_array_elements(data -> 'records') AS elem
                                            FROM company_data
                                            WHERE company_id = :company_id
                                        )
                                        ```
                                    * **Multiple CTEs:** Chain with commas (e.g., `WITH records_data AS (...), details_data AS (...)`).
                                    * **Naming:** Use descriptive CTE names based on the array key (e.g., `records_data`).
                                
                                2.  **Accessing Fields:**
                                    * Use `->>` to extract fields from the unnested `elem` alias (e.g., `elem ->> 'name'`).
                                    * Assign clear aliases in the final `SELECT` (e.g., `elem ->> 'name' AS name`).
                                
                                3.  **Casting:**
                                    * **Identifiers/Counts:** `(elem ->> 'id')::FLOAT::INTEGER AS id` (preferred for "123.0"). Use `(elem ->> 'id')::INTEGER` only if always a clean integer string.
                                    * **Numeric Fields:** Cast to `FLOAT` (e.g., `(elem ->> 'value1')::FLOAT`).
                                    * **Date Fields:** Cast to `DATE` (e.g., `(elem ->> 'date')::DATE`).
                                    * **Requirement:** Apply casts *before* comparisons, joins, or aggregations.
                                
                                4.  **Division by Zero:**
                                    * Use `NULLIF(denominator, 0)` for safe division: `(elem ->> 'value1')::FLOAT / NULLIF((elem ->> 'value2')::FLOAT, 0)`.
                                
                                5.  **Joining Arrays:**
                                    * Create separate CTEs for each array (e.g., `records_data`, `details_data`).
                                    * Use standard SQL `JOIN` (`INNER` by default, `LEFT` if explicitly required for missing matches).
                                    * Join on extracted and casted identifier fields: `(records_data.elem ->> 'id')::FLOAT::INTEGER = (details_data.elem ->> 'id')::FLOAT::INTEGER`.
                                    * **Prohibited:** Do not use `FULL OUTER JOIN`.
                                
                                6.  **Aggregation:**
                                    * Use standard SQL aggregates (`SUM`, `AVG`, `COUNT`, etc.) on casted fields.
                                    * For aggregation tasks, create a CTE to compute aggregates:
                                        Example:
                                            WITH records_agg AS (
                                                SELECT
                                                    (elem ->> 'id')::FLOAT::INTEGER AS id,
                                                    SUM((elem ->> 'value1')::FLOAT) AS total_value
                                                FROM company_data,
                                                    jsonb_array_elements(data -> 'records') AS elem
                                                WHERE company_id = :company_id
                                                GROUP BY (elem ->> 'id')::FLOAT::INTEGER
                                            )
                                    * Use final `SELECT` aliases in `GROUP BY` and `ORDER BY` clauses.
                                
                                7.  **NULL Handling:**
                                    * Use `COALESCE(field, 0)` for numeric fields to replace `NULL`s with `0` in aggregations.
                                    * **Percentage Calculations:**
                                        * Formula: `(value / NULLIF(denominator, 0)) * 100`.
                                        * Ensure non-`NULL` results: `COALESCE((GREATEST(value, 0) / NULLIF(denominator, 0)) * 100, 0)`.
                                
                                8.  **Total Row (If Requested):**
                                    * Compute totals for numeric columns (using `SUM`) and averages (using `AVG`).
                                    * Store the main report in a CTE named `report_data`.
                                    * `UNION ALL` with a `totals` CTE:
                                        Example:
                                            WITH report_data AS (...),
                                            totals AS (
                                                SELECT
                                                    'Total' AS name,
                                                    SUM(numeric_column) AS numeric_column,
                                                    AVG(avg_column) AS avg_column
                                                FROM report_data
                                            )
                                            SELECT * FROM report_data
                                            UNION ALL
                                            SELECT * FROM totals
                                
                                9.  **Column Naming:**
                                    * Use clear, human-readable aliases (e.g., `Total Value`).
                                    * For sums, use the field name directly (e.g., `Value1`).
                                    * For averages, prefix with `Avg` (e.g., `Avg Value`).
                                    * Exclude identifier fields (e.g., `id`) from the final `SELECT` unless explicitly requested.
                                
                                10. **Profit Calculations (If Specified):**
                                    * **Gross Profit:** `value1 - value2` (or specified fields).
                                    * **Gross Profit Percentage:** `COALESCE((GREATEST(value1 - value2, 0) / NULLIF(value1, 0)) * 100, 0)`.
                                
                                ### **Query Structure Checklist**
                                
                                * Starts with a `WITH` clause defining CTEs for unnesting.
                                * Each CTE includes `WHERE company_id = :company_id`.
                                * Uses `INNER JOIN` by default, `LEFT JOIN` only when required.
                                * Employs `COALESCE` and `NULLIF` for safe calculations.
                                * Stores report in `report_data` CTE and `UNION`s with a total row (if requested).
                                * Syntactically correct (no missing commas, correct parentheses, valid CTE structure).
                                * **Prohibited:** Do not use column aliases in `WHERE` clauses.
                                * **Prohibited:** Do not include `ORDER BY` in the final query if a total row is added.
                                * Fetches all required data; does not assume `NULL` or `0` values unless specified.
                                * Includes identifier fields in CTEs for joins but excludes from final `SELECT` unless requested.
                                
                                ### **Task Details**
                                * **Task Description:** `{{{{TASK_DESCRIPTION_PLACEHOLDER}}}}`
                                * **Required Data Summary:** `{{{{REQUIRED_DATA_PLACEHOLDER}}}}`
                                * **Company ID:** `{{{{COMPANY_ID_PLACEHOLDER}}}}`
                                * **Parameters:** Use `:company_id` and `:to_date` in the query.
                                
                                ### **Output Format**
                                * Output **only** the raw PostgreSQL query.
                                * **No comments, explanations, or markdown (e.g., ```sql).**
                                * The query must begin with a `WITH` clause.
                                * Verify syntax before outputting.
                                
                                ---"""
        # --- *** End Modified SQL Instruction *** ---

        try:
            # Initialize SQL model with the new system instruction
            sql_gemini = initialize_gemini_model(
                system_instruction=sql_instruction, model_name="gemini-2.5-pro-preview-06-05"
            )  # Or a more powerful model if needed
            logger.debug("SQL Gemini model initialized for JSONB querying.")
        except Exception as model_init_error:
            logger.error(
                f"❌ Failed to initialize SQL Gemini model: {model_init_error}",
                exc_info=True,
            )
            return [
                {
                    "type": "text",
                    "data": f"Error: Failed to initialize the SQL generation component. {model_init_error}",
                }
            ]

        # --- Loop Through Tasks ---
        for i, task in enumerate(tasks):
            task_description = task.get("description", f"Unnamed Task {i + 1}")
            task_type = task.get("task_type")
            required_data = task.get(
                "required_data_summary", "No data summary provided"
            )
            viz_type = task.get("visualization_type")

            logger.info(
                f"\n  Task {i + 1}/{len(tasks)}: '{task_description}' ({task_type})"
            )

            if not task_type or not required_data:
                logger.warning(f"  [!] Skipping task - missing type or data summary")
                response = {
                        "type": "text",
                        "data": f"Skipping sub-task '{task_description}' due to incomplete definition from AI.",
                    }
                results.append(
                    response
                )
                insert_message_into_db({
                    "chat_id": chat_id,
                    "sender": "System",
                    "result": json.dumps(response)
                })
                continue

            try:
                # --- Step 3a: Generate SQL Query using AI ---
                logger.info(f"    Generating SQL for JSONB...")
                if sql_gemini is None:  # Should not happen if initialization succeeded
                    logger.error("   [✗] SQL Gemini model not initialized!")
                    response = {
                            "type": "text",
                            "data": f"Error processing task '{task_description}': SQL generation component not ready.",
                        }
                    results.append(
                        response
                    )
                    insert_message_into_db({
                        "chat_id": chat_id,
                        "sender": "System",
                        "result": json.dumps(response)
                    })
                    continue

                # --- Construct the specific prompt for this task ---
                sql_prompt = (
                    f"Task Description: {task_description}\n"
                    f"Required Data Summary: {required_data}\n"
                    f"Company ID for Query: {company_id}\n"  # Inject the actual company_id
                    f"Generate the PostgreSQL query using ONLY the provided schema, column Descriptions, rules and adhering strictly to the JSONB "
                    f"querying rules, including the :company_id parameter and correct field access in "
                    f"SELECT/GROUP "
                    f"BY/ORDER BY."
                )  # Added reminder

                sql_chat = sql_gemini.start_chat()
                sql_response = sql_chat.send_message(sql_prompt)
                sql_query_text = clean_response_text(sql_response.text)
                logger.info(
                    f"    Generated SQL:\n{sql_query_text}"
                )  # Log the generated SQL

                # --- Basic SQL Validation ---
                stripped_sql = sql_query_text.lower().strip()
                if not sql_query_text or not (
                        stripped_sql.startswith("select") or stripped_sql.startswith("with")
                ):
                    logger.warning(
                        f"    [✗] Invalid or empty SQL query generated by AI (must start with SELECT or WITH): '"
                        f"{sql_query_text}'"
                    )
                    response = {
                            "type": "text",
                            "data": f"Could not generate a valid SQL query (must start with SELECT or WITH) for task: "
                                    f"'{task_description}'. AI Output: '{sql_query_text}'",
                        }
                    results.append(
                        response
                    )
                    insert_message_into_db({
                        "chat_id": chat_id,
                        "sender": "System",
                        "result": json.dumps(response)
                    })
                    continue

                if ":company_id" not in sql_query_text:
                    logger.warning(
                        f"    [✗] Generated SQL query is missing the mandatory ':company_id' para"
                        f"meter: '{sql_query_text}'"
                    )
                    response = {
                            "type": "text",
                            "data": f"Generated SQL query is invalid (missing ':company_id') for task: '"
                                    f"{task_description}'. Cannot execute safely.",
                        }
                    results.append(
                        response
                    )
                    insert_message_into_db({
                        "chat_id": chat_id,
                        "sender": "System",
                        "result": json.dumps(response)
                    })
                    continue

                logger.info(f"    [✓] SQL query generated and basic validation passed.")

                # --- Step 3b: Fetch Data from Database ---
                logger.info(f"    Fetching data using JSONB query...")
                # --- Use the new function with parameter binding ---
                data = sql_query_with_params(
                    sql_query_text,
                    params={"company_id": company_id},
                )

                if not data:
                    # It's possible the query is correct but returns no data matching criteria
                    logger.info(
                        f"    [!] Query executed successfully but returned no data."
                    )
                    response = {
                            "type": "text",
                            "data": f"For '{task_description}': The query executed successfully but found no matching "
                                    f"data for Company ID {company_id} based on the criteria.",
                        }
                    results.append(
                        response
                    )
                    insert_message_into_db({
                        "chat_id": chat_id,
                        "sender": "System",
                        "result": json.dumps(response)
                    })
                    continue
                else:
                    logger.info(f"    [✓] Fetched {len(data)} records")

                # --- Step 3c: Generate Insight, Visualization, or Report ---
                # (No changes needed in this section, it processes the fetched 'data')

                # (Insight Generation Logic)
                if task_type == "insight":
                    logger.info(f"    Generating insight...")
                    if insight_gemini is None:
                        insight_instruction = """You are an analyst. Based on the provided data (in JSON format) and 
                        the original request, generate a concise textual insight.
                        - Focus on answering the specific question asked in the 'Original Request'.
                        - Be factual and base your answer ONLY on the provided data.
                        - Keep the insight brief (1-3 sentences).
                        - Output ONLY the insight text. No extra formatting or greetings."""
                        try:
                            insight_gemini = initialize_gemini_model(
                                model_name="gemini-1.5-flash",
                                system_instruction=insight_instruction,
                            )
                            logger.debug("Insight Gemini model initialized.")
                        except Exception as model_init_error:
                            logger.error(
                                f"   [✗] Failed to initialize Insight Gemini model: {model_init_error}",
                                exc_info=True,
                            )
                            response = {
                                    "type": "text",
                                    "data": f"Error processing task '{task_description}': Insight generation "
                                            f"component failed to initialize.",
                                }
                            results.append(
                                response
                            )
                            insert_message_into_db({
                                "chat_id": chat_id,
                                "sender": "System",
                                "result": json.dumps(response)
                            })
                            continue

                    insight_prompt = f"""
                    Data (JSON format):
                    {json.dumps(data, indent=2, default=str)}

                    Original Request for this Insight:
                    "{task_description}"

                    Generate the insight based *only* on the data provided:
                    """
                    insight_chat = insight_gemini.start_chat()
                    insight_response = insight_chat.send_message(insight_prompt)
                    insight_text = clean_response_text(insight_response.text)
                    logger.debug(f"Generated Insight: {insight_text}")
                    logger.info(f"    [✓] Insight generated")
                    response = {"type": "text", "data": insight_text}
                    results.append(response)
                    insert_message_into_db({
                        "chat_id": chat_id,
                        "sender": "System",
                        "result": json.dumps(response)
                    })

                # (Visualization Generation Logic)
                elif task_type == "visualization":
                    viz_type_str = viz_type if viz_type else "chart"
                    logger.info(f"    Generating {viz_type_str} visualization...")

                    if not viz_type or viz_type not in ["bar", "line"]:
                        logger.warning(
                            f"    [!] Invalid or missing visualization type '{viz_type}' specified for task."
                        )
                        response = {
                                "type": "text",
                                "data": f"Skipping visualization for '{task_description}': Invalid or missing chart"
                                        f" type ('{viz_type}'). Requires 'bar' or 'line'.",
                            }
                        results.append(
                            response
                        )
                        insert_message_into_db({
                            "chat_id": chat_id,
                            "sender": "System",
                            "result": json.dumps(response)
                        })
                        continue

                    if plotly_gemini is None:
                        plotly_instruction = f""" You are a data visualization expert using Plotly.js. Given a 
                        dataset (as a JSON list of objects), a description of the desired visualization, 
                        and the required chart type (bar or line), generate the Plotly JSON configuration (
                        specifically the 'data' and 'layout' objects).

                        Rules:
                        - Create a meaningful title for the chart based on the description. Use the exact column 
                        names (keys) from the dataset for 'x' and 'y' keys in the data trace(s).
                        - Ensure the generated JSON is syntactically correct and contains ONLY the 'data' (list) and 
                        'layout' (object) keys at the top level.
                        - Map the data fields appropriately to x and y axes based on the description and chart type (
                        'bar' or 'line'). Infer appropriate axes labels from the data keys if not obvious.
                        - Generate ALL necessary fields for a basic, valid Plotly chart (e.g., 'type', 'x', 
                        'y' in trace; 'title' in layout). Add axis titles ('xaxis': {{"title": "X Label"}}, 
                        'yaxis': {{"title": "Y Label"}}).
                        - If multiple traces are needed (e.g., comparing two values per category), generate a list of 
                        trace objects within the 'data' list.
                        - ONLY output the JSON object starting with `{{` and ending with `}}`.
                        - Do not include any explanations, comments, code blocks (like ```json), or other text.

                        Example Output Format:
                        {{
                          "data": [
                            {{
                              "x": [/* array of x-values */],
                              "y": [/* array of y-values */],
                              "type": "{viz_type}",
                              "name": "Optional Trace Name"
                            }}
                           ],
                          "layout": {{
                            "title": "Chart Title Based on Description",
                            "xaxis": {{"title": "X-Axis Label"}},
                            "yaxis": {{"title": "Y-Axis Label"}}
                          }}
                        }}
                        """
                        try:
                            plotly_gemini = initialize_gemini_model(
                                system_instruction=plotly_instruction
                            )
                            logger.debug("Plotly Gemini model initialized.")
                        except Exception as model_init_error:
                            logger.error(
                                f"   [✗] Failed to initialize Plotly Gemini model: {model_init_error}",
                                exc_info=True,
                            )
                            response = {
                                    "type": "text",
                                    "data": f"Error processing task '{task_description}': Visualization generation "
                                            f"component failed to initialize.",
                                }
                            results.append(
                                response
                            )
                            insert_message_into_db({
                                "chat_id": chat_id,
                                "sender": "System",
                                "result": json.dumps(response)
                            })
                            continue

                    data_keys = (
                        list(data[0].keys()) if data else []
                    )  # Get keys from first record
                    plotly_prompt = f"""
                    Dataset (JSON format, keys available: {data_keys}):
                    {json.dumps(data, indent=2, default=str)}

                    Visualization Description:
                    "{task_description}"

                    Required Chart Type:
                    "{viz_type}"

                    Generate the Plotly JSON configuration ('data' and 'layout' objects only):
                    """
                    plotly_chat = plotly_gemini.start_chat()
                    plotly_response = plotly_chat.send_message(plotly_prompt)
                    plotly_json_text = clean_response_text(plotly_response.text)
                    logger.debug(f"Raw Plotly JSON response: {plotly_response.text}")
                    logger.debug(f"Cleaned Plotly JSON response: {plotly_json_text}")

                    try:
                        # More robust check for valid JSON object string
                        if not (
                                plotly_json_text.startswith("{")
                                and plotly_json_text.endswith("}")
                        ):
                            # Try removing potential leading/trailing garbage if simple cleaning failed
                            match = re.search(r"\{.*}", plotly_json_text, re.DOTALL)
                            if match:
                                plotly_json_text = match.group(0)
                            else:
                                raise ValueError(
                                    "Plotly response is not a valid JSON object string."
                                )

                        plotly_json = json.loads(plotly_json_text)
                        # Basic validation
                        if (
                                not isinstance(plotly_json, dict)
                                or "data" not in plotly_json
                                or "layout" not in plotly_json
                        ):
                            raise ValueError(
                                "Plotly JSON missing 'data' or 'layout' key at the top level, or is not an object."
                            )
                        if not isinstance(plotly_json["data"], list):
                            raise ValueError("Plotly 'data' key must be a list.")
                        if not isinstance(plotly_json["layout"], dict):
                            raise ValueError("Plotly 'layout' key must be an object.")
                        # Optional: Deeper validation of trace structure if needed

                        logger.info(f"    [✓] Visualization ({viz_type}) generated")
                        response = {"type": "graph", "data": plotly_json}
                        results.append(response)
                    except (json.JSONDecodeError, ValueError) as e:
                        logger.error(
                            f"    [✗] Failed to parse or validate Plotly JSON: {e}",
                            exc_info=False,
                        )  # Keep exc_info=False
                        logger.error(
                            f"    Problematic Plotly JSON text: {plotly_json_text}"
                        )
                        response =                             {
                                "type": "text",
                                "data": f"Error generating visualization for '{task_description}': Invalid Plotly "
                                        f"configuration received from AI. Details: {e}",
                            }
                        results.append(

                        )
                        insert_message_into_db({
                            "chat_id": chat_id,
                            "sender": "System",
                            "result": json.dumps(response)
                        })

                # (Report Generation Logic)
                elif task_type == "report":
                    logger.info(f"    Generating Excel report...")
                    try:
                        if (
                                not data
                        ):  # Should have been caught earlier, but double check
                            logger.warning(
                                f"    [!] No data to generate Excel report for '{task_description}'."
                            )
                            response = {
                                    "type": "text",
                                    "data": f"No data available to generate report: '{task_description}'",
                                }
                            results.append(
                                response
                            )
                            insert_message_into_db({
                                "chat_id": chat_id,
                                "sender": "System",
                                "result": json.dumps(response)
                            })
                            continue

                        df = pd.DataFrame(data)

                        excel_buffer = io.BytesIO()
                        # Use a modern engine like openpyxl
                        df.to_excel(
                            excel_buffer,
                            index=False,
                            sheet_name="ReportData",
                            engine="openpyxl",
                        )
                        excel_buffer.seek(0)

                        # Generate AI title for the filename
                        ai_generated_title = task_description  # Fallback
                        if title_gemini is None:
                            title_instruction = """You are an expert at creating concise, descriptive titles for data 
                            reports.
                                        Given a task description and optionally some of the data's column names, 
                                        generate a short (3-7 words) title suitable for a filename.
                                        The title should accurately reflect the report's content. Use underscores 
                                        instead of spaces.
                                        Output ONLY the title text. No extra formatting, explanations, or quotation 
                                        marks.
                                        Example: If task is "Report of sales per region for Q1" -> 
                                        "Q1_Sales_by_Region_Report"
                                        Example: If task is "List active users and their last login" -> 
                                        "Active_Users_Last_Login"
                                        """
                            try:
                                title_gemini = initialize_gemini_model(
                                    model_name="gemini-1.5-flash",
                                    system_instruction=title_instruction,
                                )
                                logger.debug(
                                    "Title Gemini model initialized for reports."
                                )
                            except Exception as model_init_error:
                                logger.error(
                                    f"   [✗] Failed to initialize Title Gemini model: {model_init_error}",
                                    exc_info=True,
                                )
                                # Continue with fallback title

                        if title_gemini:
                            title_prompt_parts = [
                                f'Task Description:\n"{task_description}"\n'
                            ]
                            if not df.empty:
                                title_prompt_parts.append(
                                    f"Data Columns (first few):\n{list(df.columns)[:5]}\n"
                                )
                            title_prompt_parts.append(
                                "Generate a short, filename-friendly title (3-7 words, use underscores):"
                            )
                            title_prompt = "".join(title_prompt_parts)

                            try:
                                title_chat = title_gemini.start_chat()
                                title_response = title_chat.send_message(title_prompt)
                                generated_title_text = clean_response_text(
                                    title_response.text
                                )
                                # Further clean/validate the title
                                generated_title_text = generated_title_text.replace(
                                    " ", "_"
                                )[
                                                       :100
                                                       ]  # Limit length
                                if generated_title_text and re.match(
                                        r"^\w+$", generated_title_text.replace("_", "")
                                ):
                                    ai_generated_title = generated_title_text
                                    logger.info(
                                        f"    AI Generated Title: {ai_generated_title}"
                                    )
                                else:
                                    logger.warning(
                                        f"    AI generated title was invalid or empty ('{generated_title_text}'), "
                                        f"using fallback."
                                    )
                            except Exception as title_gen_error:
                                logger.error(
                                    f"    Error generating title with AI: {title_gen_error}",
                                    exc_info=False,
                                )
                                # Continue with fallback title

                        # Create a safe filename FROM THE AI TITLE (or fallback task_description)
                        # Replace invalid chars, ensure it's not empty, truncate
                        safe_filename_base = re.sub(
                            r"[^\w-]", "_", ai_generated_title
                        ).strip("_")
                        if not safe_filename_base:
                            safe_filename_base = f"report_task_{i + 1}"
                        filename = f"{safe_filename_base[:50]}.xlsx"  # Truncate further for safety
                        wb = openpyxl.load_workbook(excel_buffer)
                        ws = wb.active

                        # Bold and center header row
                        for cell in ws[1]:
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal="center")

                        # Auto-adjust column widths
                        for col in ws.columns:
                            max_length = 0
                            column_letter = get_column_letter(col[0].column)
                            for cell in col:
                                try:
                                    value = (
                                        str(cell.value)
                                        if cell.value is not None
                                        else ""
                                    )
                                    max_length = max(max_length, len(value))
                                except:
                                    pass
                            ws.column_dimensions[column_letter].width = max_length + 2

                        # Optional: Number formatting
                        for row in ws.iter_rows(min_row=2):
                            for cell in row:
                                if isinstance(cell.value, int):
                                    cell.number_format = (
                                        "#,##0"  # For integers (no decimal)
                                    )
                                elif isinstance(cell.value, float):
                                    cell.number_format = (
                                        "#,##0.00"  # For floats (two decimals)
                                    )
                        # 🔥 Bold the last row (total row)
                        last_row_idx = ws.max_row
                        for cell in ws[last_row_idx]:
                            cell.font = Font(bold=True)

                        # Save formatted workbook back to bytes
                        formatted_buffer = io.BytesIO()
                        wb.save(formatted_buffer)
                        formatted_buffer.seek(0)
                        excel_bytes = formatted_buffer.getvalue()
                        # # Save to local directory
                        # output_path = os.path.join("output", "PM_KPI_Report.xlsx")  # Customize folder/filename
                        # os.makedirs(os.path.dirname(output_path), exist_ok=True)   # Ensure folder exists
                        # with open(output_path, "wb") as f:
                        #     f.write(excel_bytes)
                        excel_base64 = base64.b64encode(excel_bytes).decode("utf-8")

                        logger.info(
                            f"    [✓] Excel report '{filename}' prepared (base64 encoded)."
                        )
                        response = {
                                "type": "excel_file",
                                "data": {
                                    "filename": filename,
                                    "content_base64": excel_base64,
                                    "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                }
                        }
                        results.append(
                            response
                        )
                        insert_message_into_db({
                            "chat_id": chat_id,
                            "sender": "System",
                            "result": json.dumps(response)
                        })

                    except Exception as report_err:
                        logger.error(
                            f"    [✗] Failed to generate Excel report for '{task_description}': {report_err}",
                            exc_info=True,
                        )
                        results.append(
                            {
                                "type": "text",
                                "data": f"Error preparing Excel report for '{task_description}': {report_err}",
                            }
                        )

                else:
                    logger.warning(f"    [!] Unknown task type '{task_type}'")
                    response = {
                            "type": "text",
                            "data": f"Unknown task type '{task_type}' encountered for sub-task '{task_description}'. "
                                    f"Cannot process.",
                        }
                    results.append(
                        response
                    )
                    insert_message_into_db({
                        "chat_id": chat_id,
                        "sender": "System",
                        "result": json.dumps(response)
                    })

            except Exception as task_error:
                # Log the specific SQL query that failed if the error is likely SQL related
                if (
                        isinstance(task_error, Exception)
                        and "database" in str(task_error).lower()
                ):
                    logger.error(
                        f"    [✗] Database error processing task '{task_description}'. Failed "
                        f"Query:\n{sql_query_text}\nError: {task_error}",
                        exc_info=True,
                    )
                else:
                    logger.error(
                        f"    [✗] Error processing task '{task_description}': {task_error}",
                        exc_info=True,
                    )
                response = {
                        "type": "text",
                        "data": f"An error occurred while processing sub-task '{task_description}': {task_error}",
                    }
                results.append(
                    response
                )
                insert_message_into_db({
                    "chat_id": chat_id,
                    "sender": "System",
                    "result": json.dumps(response)
                })
            # End of task processing try-except block
        # End of loop through tasks

    except Exception as e:
        logger.error(
            f"❌ An unexpected error occurred during the main processing pipeline: {e}",
            exc_info=True,
        )
        # Append a generic error message to results if appropriate
        response = {
                "type": "text",
                "data": f"An unexpected error occurred during processing: {e}",
            }
        results.append(
            response
        )
        insert_message_into_db({
            "chat_id": chat_id,
            "sender": "System",
            "result": json.dumps(response)
        })
    # End of main try-except block

    logger.info("\n🏁 PIPELINE EXECUTION COMPLETE")
    if results:
        logger.info(f"Returning {len(results)} result items.")
        # Log types of results generated
        result_types = [r.get("type", "unknown") for r in results]
        logger.info(f"Result types: {', '.join(result_types)}")
    else:
        logger.info("No results were generated.")
    return results
