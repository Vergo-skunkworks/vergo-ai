import os
from datetime import datetime
import pandas as pd
from sqlalchemy import create_engine, text
import json
from contextlib import contextmanager
import re
import logging
import google.generativeai as genai
from dotenv import load_dotenv
import ast
import uuid
from typing import List, Dict, Any, Union
import io
import base64
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any

# Ensure this is loaded at the very beginning
load_dotenv()

log_format = "%(levelname)s: %(message)s"
logging.basicConfig(level=logging.INFO, format=log_format)
logger = logging.getLogger(__name__)


# --- Database Connection (from updatedIntegrations.py, as before) ---
@contextmanager
def get_db_connection():
    """Provides a managed database connection."""
    db_name = os.environ.get("DB_NAME")
    db_user = os.environ.get("DB_USER")
    db_password = os.environ.get("DB_PASSWORD")
    instance_connection_name = os.environ.get("INSTANCE_CONNECTION_NAME")

    if not all([db_name, db_user, db_password, instance_connection_name]):
        missing_vars = []
        if not db_name: missing_vars.append("DB_NAME")
        if not db_user: missing_vars.append("DB_USER")
        if not db_password: missing_vars.append("DB_PASSWORD")
        if not instance_connection_name: missing_vars.append("INSTANCE_CONNECTION_NAME")
        raise ValueError(f"Missing required environment variables: {', '.join(missing_vars)}")

    from urllib.parse import quote_plus

    db_host = os.environ.get("DB_HOST", "35.190.189.103")

    # conn_string = (
    #     f"postgresql+psycopg2://{db_user}:{quote_plus(db_password)}@/{db_name}"
    #     f"?host={db_host}"
    # )

    # # Alternative format (both should work):
    conn_string = (
        f"postgresql+psycopg2://{db_user}:{quote_plus(db_password)}@/{db_name}"
        f"?host=/cloudsql/{instance_connection_name}"
    )

    engine = None
    conn = None
    try:
        logger.debug("Attempting to connect to the database.")
        engine = create_engine(
            conn_string,
            pool_size=5,
            max_overflow=10,
            pool_timeout=30,
            pool_recycle=1800,
        )
        conn = engine.connect()
        logger.debug("Database connection successful.")
        yield conn, engine
    except Exception as e:
        logger.error(f"âŒ Database connection failed: {e}", exc_info=True)
        raise
    finally:
        if conn:
            conn.close()
            logger.debug("Database connection closed.")
        if engine:
            engine.dispose()
            logger.debug("Database engine disposed.")


def sql_query_with_params(query: str, params: dict = None) -> List[Dict[str, Any]]:
    """
    Runs a SQL SELECT query on the PostgreSQL database with parameters
    and returns results as a list of dictionaries.
    """
    logger.debug(f"Executing SQL query: {query} with params: {params}")
    try:
        with get_db_connection() as (conn, engine):
            result = pd.read_sql_query(text(query), conn, params=params)
            result = result.where(pd.notnull(result), None)
            data = result.to_dict(orient="records")
            logger.debug(f"Query executed successfully. Rows returned: {len(data)}")
            if data:
                sample_size = min(len(data), 2)
                logger.debug(
                    f"Query result sample: {json.dumps(data[:sample_size], indent=2, default=str)}"
                )
            else:
                logger.debug("Query returned no data.")
            return data
    except Exception as e:
        logger.error(f"âŒ Error executing parameterized SQL query: {e}", exc_info=True)
        raise


# --- Data Retrieval Functions (Adapted for New Schema Logic) ---

def get_file_data_and_schema(file_id: str) -> Dict[str, Any]:
    """
    Retrieves the 'data' (actual records from the file) and 'schema' for a specific file_id
    from the 'file_data' table.
    """
    logger.debug(f"Retrieving data and schema for file_id: {file_id}")
    query = text("SELECT data, schema FROM public.file_data WHERE file_id = :file_id LIMIT 1")

    try:
        with get_db_connection() as (conn, engine):
            result = conn.execute(query, {"file_id": file_id}).fetchone()

            if result:
                file_data_content = result[0]
                file_schema_content = result[1]

                file_data_content = file_data_content if file_data_content is not None else []
                file_schema_content = file_schema_content if file_schema_content is not None else {}

                logger.debug(f"Data and schema retrieved successfully for file_id {file_id}.")
                return {
                    "data": file_data_content,
                    "schema": file_schema_content
                }
            else:
                logger.warning(f"No data or schema found for file_id: {file_id}")
                return {"data": [], "schema": {}}

    except Exception as e:
        logger.error(
            f"âŒ Error retrieving data and schema for file_id {file_id}: {e}",
            exc_info=True,
        )
        raise


# --- LLM Initialization (No changes) ---
def initialize_gemini_model(model_name="gemini-2.0-flash", system_instruction=None):
    """Initializes and configures the Gemini model."""
    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        raise ValueError("GEMINI_API_KEY must be set in environment variables.")

    genai.configure(api_key=api_key)

    generation_config = {
        "temperature": 0.0,
        "top_p": 0.95,
        "top_k": 64,
        "max_output_tokens": 8192,
        "response_mime_type": "text/plain",
    }
    safety_settings = [
        {"category": "HARM_CATEGORY_HARASSMENT", "threshold": "BLOCK_MEDIUM_AND_ABOVE"},
        {
            "category": "HARM_CATEGORY_HATE_SPEECH",
            "threshold": "BLOCK_MEDIUM_AND_ABOVE",
        },
        {
            "category": "HARM_CATEGORY_SEXUALLY_EXPLICIT",
            "threshold": "BLOCK_MEDIUM_AND_ABOVE",
        },
        {
            "category": "HARM_CATEGORY_DANGEROUS_CONTENT",
            "threshold": "BLOCK_MEDIUM_AND_ABOVE",
        },
    ]

    model_kwargs = {
        "model_name": model_name,
        "safety_settings": safety_settings,
        "generation_config": generation_config,
    }
    if system_instruction:
        model_kwargs["system_instruction"] = system_instruction

    try:
        model = genai.GenerativeModel(**model_kwargs)
        logger.debug(f"Gemini model '{model_name}' initialized successfully.")
        return model
    except Exception as e:
        logger.error(
            f"âŒ Failed to initialize Gemini model '{model_name}': {e}", exc_info=True
        )
        raise


def clean_response_text(text):
    """Removes markdown code blocks and trims whitespace."""
    text = re.sub(r"^```[a-zA-Z]*\s*|\s*```$", "", text, flags=re.MULTILINE)
    return text.strip()


def parse_llm_json_response(response_text: str) -> Dict[str, Any]:
    """
    Safely parses an LLM's JSON response, handling common LLM output quirks.
    Expects a single JSON object.
    """
    # Remove markdown code blocks if present
    cleaned_text = clean_response_text(response_text)

    # Try to find a JSON object within the string if the LLM adds extra text
    match = re.search(r"\{.*\}", cleaned_text, re.DOTALL)
    if match:
        json_str = match.group(0)
    else:
        # If no explicit JSON object found, assume the whole string *should* be JSON
        json_str = cleaned_text

    try:
        # ast.literal_eval is safer for Python objects but requires converting JS booleans/nulls
        # json.loads is direct for JSON
        return json.loads(json_str)
    except json.JSONDecodeError as e_json:
        logger.warning(f"JSON parsing failed: {e_json}. Attempting AST literal_eval.")
        try:
            # Fallback to ast.literal_eval for more flexible parsing, converting common JSON primitives
            python_compatible = (
                json_str.replace("null", "None")
                .replace("true", "True")
                .replace("false", "False")
            )
            parsed = ast.literal_eval(python_compatible)
            if isinstance(parsed, dict):
                return parsed
            else:
                raise ValueError(f"AST eval did not result in a dictionary: {parsed}")
        except (SyntaxError, ValueError, TypeError) as e_ast:
            logger.error(
                f"Failed to parse LLM JSON response. Raw response: '{response_text}'. "
                f"JSON Error: {e_json}. AST Error: {e_ast}",
                exc_info=False,
            )
            raise ValueError(
                f"Failed to parse LLM response as JSON. Check AI response format. Raw: '{response_text[:200]}...'"
            )
def parse_tasks_response(response_text):
    """
    Safely parse the task decomposition response into a Python list of dictionaries.
    Handles JSON null values by converting them to Python None.
    """
    try:
        # First try ast.literal_eval after replacing null/true/false
        python_compatible = (
            response_text.replace("null", "None")
            .replace("true", "True")
            .replace("false", "False")
        )
        tasks = ast.literal_eval(python_compatible)
        if not isinstance(tasks, list):
            raise ValueError("Decomposition did not return a list (evaluated by ast).")
        # Basic validation of task structure
        for task in tasks:
            if not isinstance(task, dict):
                raise ValueError(f"Task item is not a dictionary: {task}")
            if "task_type" not in task or "description" not in task:
                logger.warning(f"Task missing 'task_type' or 'description': {task}")
        return tasks
    except (SyntaxError, ValueError, TypeError) as ast_error:
        logger.warning(f"AST parsing failed: {ast_error}. Trying JSON parsing.")
        try:
            # If ast fails, try json.loads
            tasks = json.loads(response_text)
            if not isinstance(tasks, list):
                raise ValueError(
                    "Decomposition did not return a list (evaluated by json)."
                )
            # Basic validation of task structure (redundant but safe)
            for task in tasks:
                if not isinstance(task, dict):
                    raise ValueError(f"Task item is not a dictionary: {task}")
                if "task_type" not in task or "description" not in task:
                    logger.warning(f"Task missing 'task_type' or 'description': {task}")
            return tasks
        except (json.JSONDecodeError, ValueError, TypeError) as json_error:
            logger.error(
                f"Failed to parse task list. Raw response: '{response_text}'. AST error: {ast_error}. "
                f"JSON error: {json_error}",
                exc_info=False,
            )  # Keep exc_info=False for cleaner logs unless debugging heavily
            raise ValueError(
                f"Failed to parse task list. Check AI response format. Raw response snippet: '{response_text[:200]}...'"
            )


def update_chat_last_updated_at(chat_id: str):
    """Updates the last_updated_at timestamp for a chat."""
    query = """
        UPDATE chats
        SET last_updated_at = CURRENT_TIMESTAMP
        WHERE chat_id = :chat_id;
    """
    params = {"chat_id": chat_id}
    with get_db_connection() as (conn, engine):
        conn.execute(text(query), params)
        conn.commit()

def insert_message_into_db(message_data: Dict[str, Any]) -> Dict[str, Any]:
    """Inserts a new message record into the 'messages' table."""
    query = """
        INSERT INTO messages (
            chat_id, sender, message_text, timestamp
        ) VALUES (
            :chat_id, :sender, :message_text, :timestamp
        ) RETURNING message_id;
    """

    # --- Prepare message_text as a string from a JSON object array ---
    # This is the part that converts your array of JSON objects into a string.
    # It checks if report_data or graph_data is provided and is a list/dict (JSON-like)
    # If so, it stringifies it for message_text.

    # If your LLM's 'message_text' field already contains the desired string, use that directly.
    # If not, and you want to stringify 'report_data' or 'graph_data' for 'message_text':
    _message_text = message_data.get("result", "")  # Start with what's provided

    logger.info("Inserting response into database")

    params = {
        "chat_id": message_data["chat_id"],
        "sender": message_data["sender"],
        "message_text": _message_text,  # Use the stringified version or original message_text
        "timestamp": message_data.get("timestamp", datetime.now().isoformat()),
    }

    with get_db_connection() as (conn, engine):
        result = conn.execute(text(query), params).fetchone()
        conn.commit()
        return dict(result._mapping) if result else None


def remove_newlines(text: str) -> str:
    """
    Removes all newline characters from the given text.

    Parameters
    ----------
    text : str
        The input text with potential newlines.

    Returns
    -------
    str
        The text with all `\n` characters removed.
    """
    return text.replace('\n', '').strip()


def get_messages_from_db(chat_id: str) -> List[Dict[str, Any]]:
    """Retrieves all messages for a given chat_id from the 'messages' table."""
    query = """
        SELECT message_id, chat_id, sender, message_text, timestamp
        FROM messages
        WHERE chat_id = :chat_id
        ORDER BY timestamp ASC;
    """
    params = {"chat_id": chat_id}
    with get_db_connection() as (conn, engine):
        result = conn.execute(text(query), params).fetchall()
        # Convert row._mapping to dict explicitly to ensure JSONB fields are handled correctly.
        # Psycopg2 should handle JSONB -> dict automatically, but good to be aware.
        return [dict(row._mapping) for row in result]


def format_chat_history_for_gemini(chat_messages: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Formats a list of database chat messages into a history list suitable for Gemini's ChatSession.
    For Gemini API, 'parts' should be a list of Content objects (often just strings).
    """
    gemini_history = []
    chat_messages = chat_messages[-6:]
    for msg in chat_messages:
        role = "user" if msg["sender"].lower() == "user" else "model"

        # 'parts' needs to be a list, even for a single text part
        parts_list = None

        if msg.get("message_text"):
            parts_list = msg["message_text"]


        # Only append a turn if there's actual content to display to the LLM
        if parts_list:
            gemini_history.append({"role": role, "parts": parts_list})

    return gemini_history

def get_category_schemas_and_description(
        company_id: int,
        selected_datasource: List[str],
        selected_datasource_type: str
) -> Dict[str, Any]:
    """
    Builds and executes a schema retrieval query for selected categories.

    Returns:
        A dict mapping category_name to its schema.
    """
    if not selected_datasource:
        return {}
    in_clause = tuple(selected_datasource)
    # in_clause = ', '.join([f"'{cat}'" for cat in selected_datasource])
    print(in_clause)

    query_forFiles = f"""
        WITH LatestFilesPerCategory AS (
            SELECT
                f.category_name,
                f.original_file_name,
                f.file_id
            FROM public.files f
            WHERE f.company_id = :company_id AND f.file_id IN {in_clause}
        )
        SELECT
            original_file_name,
            lfp.file_id,
            lfp.category_name,
            fd.schema,
            fd.description
        FROM LatestFilesPerCategory lfp
        JOIN public.file_data fd ON lfp.file_id = fd.file_id
    """

    query_forCategories = f"""
        WITH LatestFilesPerCategory AS (
            SELECT
                f.category_name,
                f.file_id,
                ROW_NUMBER() OVER (
                    PARTITION BY f.category_name
                    ORDER BY f.uploaded_at DESC, f.file_id DESC
                ) as rn
            FROM public.files f
            WHERE f.company_id = :company_id AND f.category_name IN {in_clause}
        )
        SELECT
            lfp.category_name,
            fd.schema,
            fd.description
        FROM LatestFilesPerCategory lfp
        JOIN public.file_data fd ON lfp.file_id = fd.file_id
        WHERE lfp.rn = 1;
    """

    logger.info(f"[âœ“] Running schema retrieval query for categories: {selected_datasource}")

    try:
        if selected_datasource_type == "files":
            query = query_forFiles
        else:
            query = query_forCategories
        rows = sql_query_with_params(query, params={"company_id": company_id})
        if selected_datasource_type == "files":
            return {
                row["file_id"]: {
                    "file_name": row["original_file_name"],
                    "schema": row["schema"],
                    "description": row["description"]
                }
                for row in rows
            }
        else:
            return {
                row["category_name"]: {
                    "schema": row["schema"],
                    "description": row["description"]
                }
                for row in rows
            }
    except Exception as e:
        logger.exception(f"Failed to retrieve category schemas: {str(e)}")
        return {}


def generate_insight(data, task_description, insight_gemini=None):
    logger.info("    Generating insight...")

    if insight_gemini is None:
        insight_instruction = f"""**Your Role and Goal:**
                                    You are an expert data analyst. Your primary goal is to interpret raw data and transform it into a concise, easy-to-understand narrative for a business audience. You must go beyond the numbers to explain what they *mean*.
                                    
                                    **Your Task:**
                                    Based on the provided JSON data and the 'Original Request' below, you will generate a single, self-contained HTML snippet that presents a clear and descriptive textual insight.
                                    
                                    **Critical Analysis Requirements:**
                                    - **Analyze and Interpret:** Do not just list the data. Identify significant trends, patterns, outliers, and relationships within the data.
                                    - **Explain the "Why":** Provide context and potential reasons for the trends you observe. Answer the question, "So what?" for the reader.
                                    - **Synthesize Information:** Your output should be a cohesive narrative, not a list of disconnected facts. Tell a story with the data that directly answers the user's original request.
                                    - **Be Factual:** Base your entire analysis ONLY on the data provided. Do not invent or assume information.
                                    
                                    **Strict Formatting Rules:**
                                    - Your entire output must be a SINGLE self-contained HTML snippet.
                                    - The root element must be a `<div>` with inline CSS styling.
                                    - Use semantic tags for structure (e.g., `<h3>`, `<p>`, `<strong>`, `<ul>`, `<li>`).
                                    - The insight should be brief and impactful.
                                    - Absolutely no text, comments, or explanations should exist outside of the single root `<div>` tag.
                                    - No Border Style. and do not center the main div horizentaly as it should be display left align.
                                    - Do not add padding or Margin to main div.
                                    - Font Family should typography and size should be 16. For heading size should be relevant.
                                    - Explicitly mention margin 0 and padding 0 for main div and first text inside that div as there are previous by default margin and padding. it should be only for top component and first text compoent. but should have margin for better visisbility.
                                    - For points use <ul>. The structure should be nice and looks goood.
                                    ---
                                    
                                    ** Data is: {data}**
                                    ** Description: {task_description}
                                    """
        try:
            insight_gemini = initialize_gemini_model(
                model_name="gemini-2.0-flash",
                system_instruction=insight_instruction,
            )
            logger.debug("Insight Gemini model initialized.")
        except Exception as model_init_error:
            logger.error(
                f"   [âœ—] Failed to initialize Insight Gemini model: {model_init_error}",
                exc_info=True,
            )
            return None

    insight_prompt = f"""
    Data (JSON format):
    {json.dumps(data, indent=2, default=str)}

    Original Request for this Insight:
    "{task_description}"

    Generate the insight based *only* on the data provided:
    """

    try:
        insight_chat = insight_gemini.start_chat()
        insight_response = insight_chat.send_message(insight_prompt)
        insight_text = clean_response_text(insight_response.text)
        logger.debug(f"Generated Insight: {insight_text}")
        logger.info("    [âœ“] Insight generated")
        insight_text = remove_newlines(insight_text)
        return insight_text

    except Exception as insight_error:
        logger.error(
            f"   [âœ—] Failed to generate insight: {insight_error}",
            exc_info=True,
        )
        return None


def handle_chat_name_generation( chatId, prompt):
    """
    Handle all logic related to generating and updating chat names
    Returns: bool indicating if name was updated
    """
    try:
        conn = get_db_connection()
        model = initialize_gemini_model()

        # Check if chat exists and needs name generation
        chat_query = """SELECT name FROM chats WHERE chat_id = :chat_id"""
        chat_result = sql_query_with_params(chat_query, {"chat_id" : chatId})
        logger.info(chat_result)
        if not chat_result or chat_result[0]['name'] != "New Chat":
            return False


        # Prepare prompt for generating chat name
        name_prompt = f"""Based on the User Prompt Request, generate a very concise 2-3 word title that captures the main topic. 
        Keep it simple and descriptive. Only respond with the title, no additional text.

        Current message: {prompt}
        """

        # Generate name using Gemini
        response = model.generate_content(name_prompt)
        chat_name = clean_response_text(response.text)
        query = """UPDATE chats SET name = :name WHERE chat_id = :chat_id"""
        params = {"chat_id": chatId, "name": chat_name}
        with get_db_connection() as (conn, engine):
            conn.execute(text(query), params)
            conn.commit()
        return True

    except Exception as e:
        logger.error(f"Error in handle_chat_name_generation: {str(e)}")
        return False

def process_prompt(
        prompt: str, company_id: int, chat_id: str,selected_datasource: List[str], selected_datasource_type: str
) -> List[Dict[str, Any]]:
    """
    Processes a user prompt by inferring the relevant file categories and time windows,
    then performing analysis on the identified data.

    Args:
        prompt: The user's natural language request.
        company_id: The ID of the company.
        chat_id: The ID of the current chat for message logging.

    Returns:
        A list of result dictionaries, each containing 'type' and 'data',
        which have also been logged to the database.
    """
    logger.info("\nâœ¨ STEP 1: PROCESSING USER PROMPT")
    logger.info(f'Received Prompt: "{prompt}" for company_id: {company_id}')

    # Input validation
    if not isinstance(company_id, int) or company_id <= 0:
        error_msg = f"Error: Invalid Company ID ({company_id}). Must be a positive integer."
        logger.error(f"âŒ {error_msg}")
        insert_message_into_db(
            {"chat_id": chat_id, "sender": "llm", "message_text": error_msg, "analysis_type": "text"})
        return [{"type": "text", "data": error_msg, "message_text": error_msg}]

    logger.info(f"Target Company ID: {company_id}")

    message_history = get_messages_from_db(chat_id)
    formated_message_history = format_chat_history_for_gemini(message_history)

    response = {
        "type": "text",
        "data": prompt,
    }
    insert_message_into_db({
        "chat_id": chat_id,
        "sender": "User",
        "result": json.dumps(response)
    })

    handle_chat_name_generation(chat_id, prompt)

    results = []  # Collects the structured LLM responses to return

    try:
        logger.info(f"\nðŸ“œ STEP 2b: FETCHING SCHEMAS FOR SELECTED CATEGORIES")
        category_schemas_and_description_map = get_category_schemas_and_description(company_id, selected_datasource, selected_datasource_type)

        if not category_schemas_and_description_map:
            error_msg = f"No schemas retrieved for selected categories for company ID {company_id}."
            logger.warning(error_msg)
            llm_response_data = {"type": "text", "data": error_msg}
            results.append(llm_response_data)
            response = {"type": "text", "data": "You have not define correct data sources."}
            insert_message_into_db({"result": json.dumps(response), "chat_id": chat_id, "sender": "System"})
            return results

        logger.info(f"[âœ“] Retrieved schemas for {len(category_schemas_and_description_map)} categories.")

        # --- Step 2b: Decompose Prompt into Tasks ---
        logger.info("\nðŸ§  STEP 2b: DECOMPOSING PROMPT INTO TASKS (using JSON schema)")

        # --- DECOMPOSITION INSTRUCTION (No changes needed here) ---

        decomposition_instruction = f"""
                                    Analyze the user's request: "{prompt}"
                                    
                                    Based on this request and the Data Schema provided below, identify the specific 
                                    data analysis or reporting tasks 
                                    required to fulfill the user's objectives.
                                    **Instructions:**
                                        User Current Prompt: {prompt}. User Previous conversation already provided you. But User last prompt and llm response was: {formated_message_history[-2:]}. Follow below instructions.
                                        - Your primary focus should always be the latest prompt.
                                        - Use prior messages only when the userâ€™s current request depends on or refers to earlier context (e.g., "same as before", "add previous filters", "build on earlier report", "now add", etc.).
                                        - if user ask "now add this ___" then append these changes in previous prompt. Previous should be maintained.
                                        - If there is no clear dependency, ignore previous history and respond based only on the new prompt.
                                        - If a dependency is implied, extract the relevant information from the past messages to inform your response.
                                        - Do not repeat old information unless the user asks for it.
                                        - Be precise, and avoid redundancy.
                                    The data for these tasks resides in a single table 'file_data' within a JSONB 
                                    column named 'data'. You MUST filter 
                                    by company_id = {company_id} from files table. files table and file_data table have common file_id key
                                    The structure of this 'data' column for the relevant company is described by the 
                                    'schema'. The keys in the 'Data 
                                    Schema' (e.g., "pms", "change_order") correspond to the top-level keys within the 
                                    'data' JSONB column, each holding 
                                    an array of JSON objects.

                                    The Description about columns of this 'data' column for the relevant company is described by the 
                                    'description'. It's jsonb format where top key have description about it's columns

                                    Data Schema (Structure within the 'data' JSONB column of 'file_data') with description of columns
                                    {category_schemas_and_description_map}                         
                                    Use exact category names as given with schema as it store in database exact like this.
                                    **Guidelines for Defining Tasks:**

                                    1.  **Task Identification (Understand User Objectives):**
                                        * First, carefully read the user's prompt ("{prompt}") to understand their 
                                        primary goals or the questions they 
                                        want answered.
                                        * Identify each distinct objective. An objective might be a request for 
                                        specific information, a comparison, 
                                        a summary, a trend analysis, etc.
                                        * Each distinct objective that can be addressed with the provided schema 
                                        should correspond to a task.
                                        * User can ask multiple columns within same report. consider it single report 
                                        which specifies all columns that 
                                        user want.
                                        * Consider if data from DIFFERENT KEYS within the JSONB 'data' column (e.g., 
                                        "pms" and "change_order") needs to 
                                        be conceptually combined or related to fulfill an objective.
                                        If user only mention one task like generate insight and not mention other task like report or graph just consider one task.

                                    2.  **Task Type Assignment (`task_type`) and Refinement:**
                                        * For each objective identified in Step 1, assign a `task_type` as follows:
                                        * **If the user explicitly requests a specific task type** for an objective (
                                        e.g., "generate a report for X", 
                                        "create a chart of Y", "I need an insight about Z"):
                                            * The `task_type` for that objective **MUST** be what the user specified.
                                            * For such an explicitly typed request, generate **ONLY** that task for 
                                            that specific objective. Do not 
                                            generate alternative types (e.g., an insight if a report was asked for) 
                                            for the *same objective*.
                                        * **If the user does NOT explicitly state a task type** for an objective, 
                                        determine the most fitting type based 
                                        on the nature of the information sought:
                                            * **'insight':** Choose this if the objective is to understand *why* 
                                            something is happening, to identify key 
                                            takeaways, underlying patterns, significant trends (and their 
                                            implications), anomalies, or to get a 
                                            summarized interpretation or textual explanation that goes beyond just 
                                            presenting raw data. User prompts like 
                                            "What are the main drivers of sales?", "Explain the recent dip in 
                                            performance.", "Are there any unusual 
                                            activities?", "Summarize customer feedback trends.", "Why is X correlated 
                                            with Y?", "What does the data 
                                            suggest about Z?" often call for an 'insight'.
                                            * **'report':** Choose this if the objective is primarily to see 
                                            structured data, detailed listings, 
                                            specific values, or straightforward aggregations without requiring 
                                            significant interpretation or qualitative 
                                            explanation. User prompts like "List all overdue projects.", "Show me the 
                                            sales figures for each region.", 
                                            "What is the total number of users active last month?", "Generate a 
                                            summary table of X." typically call for a 
                                            'report'.
                                            * **'visualization':** Choose this if the objective is best met by a 
                                            graphical representation to see trends, 
                                            compare values, understand distributions, or highlight relationships 
                                            visually. User prompts like "Compare 
                                            sales over the last year.", "Show me the distribution of customer ages.", 
                                            "Plot X against Y." often call for 
                                            a 'visualization'. If so, also determine a `visualization_type` ('bar' or 
                                            'line').
                                            * **Decision Guidance:** When a prompt is somewhat ambiguous (e.g., 
                                            "Tell me about X," "Analyze Y"), 
                                            consider whether the core need is for raw/structured data ('report'), 
                                            a visual representation (
                                            'visualization'), or an interpretation/explanation/conclusion (
                                            'insight'). If the user seems to be asking a 
                                            "what," "why," or "so what" question that requires drawing conclusions 
                                            from data, lean towards 'insight'.
                                        * **Multiple Objectives:** If the user's prompt contains multiple distinct 
                                        objectives (e.g., "Give me a report on 
                                        A and also show a chart for B"), create a separate task for each objective, 
                                        applying these typing rules 
                                        individually.

                                    Based ONLY on the user prompt, the schema, and these guidelines, list the task(s).

                                    For each task, specify:
                                    1.  'task_type': 'insight', 'visualization', or 'report' (determined by Guideline)
                                    2.  'description': Brief description of the task's objective (e.g., "Report of 
                                    change orders per project manager"). 
                                    This should reflect the objective identified in Guideline 1.
                                    3.  'required_data_summary': Describe the data needed to achieve the objective, 
                                    mentioning the relevant JSON KEYS (
                                    e.g., "pms", "change_order") and the specific FIELDS from the schema (e.g., 
                                    "PM_Name from pms", "Change Orders from 
                                    change_order"). Clearly state if and how data from multiple keys needs to be 
                                    related.
                                    4.  'visualization_type': 'bar' or 'line' if `task_type` is 'visualization', 
                                    else null.

                                    Output the result as a valid Python list of dictionaries ONLY. No explanations or 
                                    markdown. Ensure keys and values 
                                    are double-quoted. Use null for missing values, not None.
                                    Example (this example shows the format; the number and nature of tasks generated 
                                    depend on the user's specific 
                                    request and the guidelines above):
                                    [
                                        {{"task_type": "report", "description": "Report linking PMs to their change 
                                        orders", "required_data_summary": 
                                        "Need PM_Name from 'pms' key and Job Number, Change Orders from 
                                        'change_order' key. Relate pms.PM_Id to 
                                        change_order.Project_Manager using extracted fields.", "visualization_type": 
                                        null}},
                                        {{"task_type": "visualization", "description": "Total change orders per PM", 
                                        "required_data_summary": "Need 
                                        PM_Name from 'pms' and Change Orders from 'change_order'. Aggregate Change 
                                        Orders grouped by PM after relating 
                                        the keys.", "visualization_type": "bar"}}
                                    ]
                                    """
        # --- End Decomposition Instruction ---

        decomposer_model = initialize_gemini_model()
        decomposer_chat = decomposer_model.start_chat(history=formated_message_history)
        response = decomposer_chat.send_message(decomposition_instruction)
        cleaned_response = clean_response_text(response.text)
        logger.debug(f"Raw task decomposition response: {response.text}")
        logger.debug(f"Cleaned task decomposition response: {cleaned_response}")

        try:
            tasks = parse_tasks_response(cleaned_response)
            if not tasks:
                logger.warning(
                    "AI task decomposition returned an empty list. No tasks to perform."
                )
                return [
                    {
                        "type": "text",
                        "data": f"I couldn't identify specific tasks from your request based on the"
                                f" available data structure for Company ID {company_id}. Could you please rephrase?",
                    }
                ]

            logger.info(f"ðŸ¤– AI identified {len(tasks)} tasks:")
            for idx, task_item in enumerate(tasks):
                logger.info(
                    f"  â€¢ Task {idx + 1}: {task_item.get('description', 'N/A')} ({task_item.get('task_type', 'N/A')})"
                )
        except (ValueError, TypeError) as e:
            logger.error(
                f"âŒ Failed to parse AI task decomposition: {e}", exc_info=False
            )
            return [
                {
                    "type": "text",
                    "data": f"Error: Could not understand the tasks required by"
                            f" the prompt. Please rephrase. (Parsing error: {e})",
                }
            ]

        # --- Step 3: Process Each Task with dynamic SQL generation ---
        logger.info("\nâš™ï¸ STEP 3: PROCESSING TASKS")
        sql_gemini = None
        plotly_gemini = None
        insight_gemini = None
        title_gemini = None

        # Initialize SQL Gemini model here once, with a dynamic system instruction
        # that describes how to query based on inferred strategies.
        # The schema part of this instruction will refer to the category_schemas_map.
        sql_instruction_for_files = f"""You are an expert PostgresSQL query writer specializing in JSONB data. Your task is to generate a single, syntactically correct PostgresSQL `SELECT` query based on the provided task, the relevant category's schema, and the specified analysis strategy.

                                        **Strictly adhere to all rules below.**
                                        **Instructions:**
                                            User Current Prompt: {prompt}. User and LLM previous conversation already provided you. But User last prompt and llm response was: {formated_message_history[-2:]}. follow below instructions.
                                            - Your primary focus should always be the latest prompt.
                                            - Use prior messages only when the userâ€™s current request depends on or refers to earlier context (e.g., "same as before", "add previous filters", "build on earlier report", "now add", etc.).
                                            - if user ask "now add this ___" then append these changes in previous prompt. Previous should be maintained.
                                            - If there is no clear dependency, ignore previous history and respond based only on the new prompt.
                                            - If a dependency is implied, extract the relevant information from the past messages to inform your response.
                                            - Do not repeat old information unless the user asks for it.
                                            - Be precise, and avoid redundancy.
                                        ### **Data Sources Overview**
                                        * Data is stored in `public.file_data` (actual records as JSONB array) and `public.files` schema is ( `uploaded_at`, `category_name`, `file_id`, `column_id` and other columns).
                                        * Each `file_data.data` column contains an array of JSON objects, representing rows from one uploaded file. Also file_data.file_id is foreign key from public.files which is files.file_id not id.
                                        * **Relevant Schema:** You will be provided with the specific schema for the file with category you are querying from. This schema describes the fields within the JSON objects in the `data` array.
                                        * **Data in public.file_data against file_id is store as jsonb like ([("date": "2025-01-06T00:00:00", "pm_id": 209012, ...)]). it just store as array of objects.
                                        * **schema:** The schema of every file with category and description of columns is: {category_schemas_and_description_map}. Also provided you description of columns. Carefully analyze description of columns so that query generation can be better and right column can be selected for query.
                                        * **company id (int)**: Use company id to filter files: {company_id}. This column is define in public.files. to get it join file_data with files table.
                                        ### **Query Strategies & Parameters**
                                        if user do not mention any time window then write query to fetch data.
                                            1. * **Query Pattern:**
                                                    ```sql
                                                    WITH unnested_data AS (
                                                        SELECT jsonb_array_elements(data) AS elem
                                                        FROM public.file_data fd
                                                        JOIN public.files f ON fd.file_id = f.file_id
                                                        WHERE and f.company_id = 1 and fd.file_id = 'something'
                                                    )
                                                    -- Rest of your query using 'elem'
                                                    ```
                                                    you have to fetch every file in separate cte with aggregation if required to get data. then final in separate cte which will make joins.
                                        **Important**: if user mention time window like analysis from january to march. Then apply filter based on data column in file. You can infer schema


                                        ### **JSONB Querying Rules (apply to `elem` from unnested data)**
                                        * **Accessing Fields:** `elem ->> 'column_name'`
                                        * **Casting:** Apply casts BEFORE operations (`::INTEGER`, `::FLOAT`, `::DATE`, `::TIMESTAMP WITH TIME ZONE`). Use `NULLIF(elem ->> 'col_name', 'NULL')::TYPE`.
                                        * For integer values, first cast as FLOAT, apply ROUND, then cast to INTEGER: ROUND((elem ->> 'column')::FLOAT)::INTEGER.
                                        * **Division by Zero:** `NULLIF(denominator, 0)`.
                                        * **Aggregation:** Standard SQL aggregates (`SUM`, `AVG`, `COUNT`).
                                        * **NULL Handling:** `COALESCE(field, 0)` for numerics.
                                        * **Important**: If you are working on column which is amount or current, it may contain currency sign. if in description column currency is available the use it or else use default dollars to remove the dollar sign and convert to int or float. e.g: SUM(COALESCE(NULLIF(REPLACE(REPLACE(elem ->> 'amount', '$', ''), ',', ''), '')::FLOAT, 0)) AS total_amount
                                        * **Total Row (If Requested):** Use `UNION ALL` with a `totals` CTE, result in `report_data` CTE.
                                        * **Column Naming:** Clear aliases. do not use _ in column nanme alias. Show proper name with every word first letter capitalized and spaces like "Total Jobs Completed or Change Order Size etc".

                                        ### **Joining Arrays:**
                                            * Create separate CTEs for each array (e.g., `records_data`, `details_data`).
                                            * Use standard SQL `JOIN` (`INNER` by default, `LEFT` if explicitly required for missing matches).
                                            * Join on extracted and casted identifier fields: `(records_data.elem ->> 'id')::FLOAT::INTEGER = (details_data.elem ->> 'id')::FLOAT::INTEGER`.
                                            * **Prohibited:** Do not use `FULL OUTER JOIN`.

                                        ### **Aggregation:**
                                            * Use standard SQL aggregates (`SUM`, `AVG`, `COUNT`, etc.) on casted fields.
                                            * For aggregation tasks, create a CTE to compute aggregates:
                                                Example:
                                                    WITH records_agg AS (
                                                        SELECT
                                                            (elem ->> 'id')::FLOAT::INTEGER AS id,
                                                            SUM((elem ->> 'value1')::FLOAT) AS total_value
                                                        FROM file_data,
                                                            jsonb_array_elements(data -> 'records') AS elem
                                                        WHERE company_id = 1
                                                        GROUP BY (elem ->> 'id')::FLOAT::INTEGER
                                                    )
                                            * Use final `SELECT` aliases in `GROUP BY` and `ORDER BY` clauses.

                                        ### **Formatting:**
                                            * If Column has value currency or amount type then must add currency sign. You MUST handle negative values correctly. The negative sign (-) must appear before the currency symbol. If you are able to detect from prompts or from columns description which currency is used in columns then used that else dollars.
                                            * Return the 'amount' value formatted with commas as thousand separators (e.g., 1,000,000) in the SQL query output.
                                            * If Column is calculating profit percentage or value percentage value, then also add % sign with value.

                                        ### **Query Structure Checklist**
                                        * Starts with `WITH`.
                                        * Applies proper casting, `COALESCE`, `NULLIF`.
                                        * Syntactically correct.
                                        * Fetches all required data; does not assume `NULL` or `0` values unless specified.
                                        * **Prohibited:** No column aliases in `WHERE`. No `ORDER BY` in final query if `UNION ALL` with total row.
                                        * Do not add extra column in reports etc. just fetch required columns.

                                        ### **Column Naming:**
                                            * Use clear, human-readable aliases (e.g., `Total Value`).
                                            * For sums, use the field name directly (e.g., `Value1`).
                                            * For averages, prefix with `Avg` (e.g., `Avg Value`).
                                            * Exclude identifier fields (e.g., `id`) from the final `SELECT` unless explicitly requested.

                                        ### **Task Details for THIS Query**
                                        * **Task Description:** `{{{{TASK_DESCRIPTION_PLACEHOLDER}}}}`
                                        * **Required Data Summary:** `{{{{REQUIRED_DATA_PLACEHOLDER}}}}`

                                        ### **Output Format**
                                        * Output **only** the raw PostgresSQL query. No comments, explanations, or markdown.
                                        * The query must begin with a `WITH` clause.
                                        * Verify syntax before outputting.
                                        ---"""

        sql_instruction_for_categories = f"""You are an expert PostgreSQL query writer specializing in JSONB data. Your task is to generate a single, syntactically correct PostgresSQL `SELECT` query based on the provided task, the relevant category's schema, and the specified analysis strategy.

                                        **Strictly adhere to all rules below.**
                                        **Instructions:**
                                            User Current Prompt: {prompt}. User and LLM previous conversation already provided you. But User last prompt and llm response was: {formated_message_history[-2:]}. follow below instructions.
                                            - Your primary focus should always be the latest prompt.
                                            - Use prior messages only when the userâ€™s current request depends on or refers to earlier context (e.g., "same as before", "add previous filters", "build on earlier report", "now add", etc.).
                                            - if user ask "now add this ___" then append these changes in previous prompt. Previous should be maintained.
                                            - If there is no clear dependency, ignore previous history and respond based only on the new prompt.
                                            - If a dependency is implied, extract the relevant information from the past messages to inform your response.
                                            - Do not repeat old information unless the user asks for it.
                                            - Be precise, and avoid redundancy.
                                        ### **Data Sources Overview**
                                        * Data is stored in `public.file_data` (actual records as JSONB array) and `public.files` (metadata like `uploaded_at`, `category_name`, `file_id`).
                                        * Each `file_data.data` column contains an array of JSON objects, representing rows from one uploaded file. Also file_data.file_id is foreign key from public.files.
                                        * **Relevant Schema:** You will be provided with the specific schema for the category you are querying from. This schema describes the fields within the JSON objects in the `data` array.
                                        * **Data in public.file_data against file_id is store as jsonb like ([("date": "2025-01-06T00:00:00", "pm_id": 209012, ...)]). it just store as array of objects.
                                        * **schema:** The schema of every file category and description of columns is: {category_schemas_and_description_map}. Also provided you description of columns. Carefully analyze description of columns so that query generation can be better and right column can be selected for query.
                                        * **company id (int)**: Use company id to filter files: {company_id}
                                        ### **Query Strategies & Parameters**
                                        if user do not mention any time window then write query to fetch data based on latest file. public.files table have uploaded_at column
                                            1.  **`latest_file` strategy:**
                                                * **Goal:** Query data from only the single latest file within the specified category.
                                                * **Filter:** Filter latest file of category by is_latest column which is boolean. it's define file is latest or not.
                                                * **Query Pattern:**
                                                    ```sql
                                                    WITH unnested_data AS (
                                                        SELECT jsonb_array_elements(data) AS elem
                                                        FROM public.file_data
                                                        WHERE is_latest = true and company_id = 1 and catefory_name = 'something'
                                                    )
                                                    -- Rest of your query using 'elem'
                                                    ```
                                                    if you have fetch every file in separate cte in get data in separate cte. then final in separate cte
                                        if user mention time window like analysis from january to march. or like other then files table have multiple files against categories. filter relevant files then union all data then do analysis.

                                            2.  **`union_all_by_time_window` strategy:**
                                                * **Goal:** Query data by combining records from ALL files within the specified `category_name` that fall within a given `start_date` and `end_date`.
                                                * **Parameters:**
                                                    * `file_ids` as `:file_ids` (a PostgreSQL array of UUID strings for an `IN` clause).
                                                    * `start_date` as `:start_date` (TEXT 'YYYY-MM-DD').
                                                    * `end_date` as `:end_date` (TEXT 'YYYY-MM-DD').
                                                * **Query Pattern:** You MUST `UNION ALL` unnested data from all files specified by `:file_ids`. Example for 2 files; extend for more in real query:
                                                    ```sql
                                                    WITH unnested_data_from_all_relevant_categories AS (
                                                        -- This CTE is provided/understood to be pre-built by Python, OR
                                                        -- LLM knows how to join file_data and files and filter by company_id, and maybe category IN (..)
                                                        SELECT jsonb_array_elements(fd.data) AS elem, f.category_name AS category_name, f.uploaded_at AS uploaded_at_file
                                                        FROM public.file_data fd
                                                        JOIN public.files f ON fd.file_id = f.file_id
                                                        WHERE f.company_id = 1 AND f.is_latest = TRUE AND f.category_name IN ([category_names_array])
                                                    )
                                                    -- Now perform joins/aggregations on elem, filtered by category_name
                                                    , pms_data_inferred AS (
                                                        SELECT (elem ->> 'pm_id')::INTEGER AS pm_id, (elem ->> 'pm_name')::TEXT AS pm_name
                                                        FROM unnested_data_from_all_relevant_categories
                                                        WHERE category_name = 'Project Managers' -- LLM infers and hardcodes filter on logical category
                                                    ),
                                                    change_orders_data_inferred AS (
                                                        SELECT (elem ->> 'pm_id')::INTEGER AS pm_id, (elem ->> 'size')::FLOAT AS size_value
                                                        FROM unnested_data_from_all_relevant_category_data
                                                        WHERE category_name = 'Change Orders' -- LLM infers and hardcodes filter on logical category
                                                    )
                                                    -- ... and then perform your JOINs and aggregations
                                                    ```
                                                    * **Ensure:** The `uploaded_at_` column is selected in the inner CTE and filtered in the main query.

                                        ### **JSONB Querying Rules (apply to `elem` from unnested data)**
                                        * **Accessing Fields:** `elem ->> 'column_name'`
                                        * **Casting:** Apply casts BEFORE operations (`::INTEGER`, `::FLOAT`, `::DATE`, `::TIMESTAMP WITH TIME ZONE`). Use `NULLIF(elem ->> 'col_name', 'NULL')::TYPE`.
                                        * For integer values, first cast as FLOAT, apply ROUND, then cast to INTEGER: ROUND((elem ->> 'column')::FLOAT)::INTEGER.
                                        * **Division by Zero:** `NULLIF(denominator, 0)`.
                                        * **Aggregation:** Standard SQL aggregates (`SUM`, `AVG`, `COUNT`).
                                        * **NULL Handling:** `COALESCE(field, 0)` for numerics.
                                        * **Total Row (If Requested):** Use `UNION ALL` with a `totals` CTE, result in `report_data` CTE.
                                        * **Column Naming:** Clear aliases.

                                        ### **Joining Arrays:**
                                            * Create separate CTEs for each array (e.g., `records_data`, `details_data`).
                                            * Use standard SQL `JOIN` (`INNER` by default, `LEFT` if explicitly required for missing matches).
                                            * Join on extracted and casted identifier fields: `(records_data.elem ->> 'id')::FLOAT::INTEGER = (details_data.elem ->> 'id')::FLOAT::INTEGER`.
                                            * **Prohibited:** Do not use `FULL OUTER JOIN`.

                                        ### **Aggregation:**
                                            * Use standard SQL aggregates (`SUM`, `AVG`, `COUNT`, etc.) on casted fields.
                                            * For aggregation tasks, create a CTE to compute aggregates:
                                                Example:
                                                    WITH records_agg AS (
                                                        SELECT
                                                            (elem ->> 'id')::FLOAT::INTEGER AS id,
                                                            SUM((elem ->> 'value1')::FLOAT) AS total_value
                                                        FROM file_data,
                                                            jsonb_array_elements(data -> 'records') AS elem
                                                        WHERE company_id = 1
                                                        GROUP BY (elem ->> 'id')::FLOAT::INTEGER
                                                    )
                                            * Use final `SELECT` aliases in `GROUP BY` and `ORDER BY` clauses.

                                        ### **Formatting:**
                                            * If Column has value currency or amount type then must add currency sign. You MUST handle negative values correctly. The negative sign (-) must appear before the currency symbol. If you are able to detect from prompts or from columns description which currency is used in columns then used that else dollars.
                                            * Return the 'amount' value formatted with commas as thousand separators (e.g., 1,000,000) in the SQL query output.
                                            * If Column is calculating profit percentage or value percentage value, then also add % sign with value.

                                        ### **Query Structure Checklist**
                                        * Starts with `WITH`.
                                        * For `latest_file` strategy: uses `WHERE is_latest=true`
                                        * For `union_all_by_time_window` strategy: constructs `UNION ALL` first, and filters using `:start_date` and `:end_date`. then join with file_data table
                                        * Applies proper casting, `COALESCE`, `NULLIF`.
                                        * Syntactically correct.
                                        * Fetches all required data; does not assume `NULL` or `0` values unless specified.
                                        * **Prohibited:** No column aliases in `WHERE`. No `ORDER BY` in final query if `UNION ALL` with total row.
                                        * Do not add extra column in reports etc. just fetch required columns.

                                        ### **Column Naming:**
                                            * Use clear, human-readable aliases (e.g., `Total Value`).
                                            * For sums, use the field name directly (e.g., `Value1`).
                                            * For averages, prefix with `Avg` (e.g., `Avg Value`).
                                            * Exclude identifier fields (e.g., `id`) from the final `SELECT` unless explicitly requested.

                                        ### **Task Details for THIS Query**
                                        * **Task Description:** `{{{{TASK_DESCRIPTION_PLACEHOLDER}}}}`
                                        * **Required Data Summary:** `{{{{REQUIRED_DATA_PLACEHOLDER}}}}`

                                        ### **Output Format**
                                        * Output **only** the raw PostgreSQL query. No comments, explanations, or markdown.
                                        * The query must begin with a `WITH` clause.
                                        * Verify syntax before outputting.
                                        ---"""

        sql_instruction = None
        if selected_datasource_type == "files":
            logger.info(f"selected datasource type is files")
            sql_instruction = sql_instruction_for_files
        else:
            sql_instruction = sql_instruction_for_categories

        try:
            sql_gemini = initialize_gemini_model(
                system_instruction=sql_instruction, model_name="gemini-2.5-pro-preview-06-05"
            )
            logger.debug("SQL Gemini model initialized for dynamic schema and strategy.")
        except Exception as model_init_error:
            error_msg = f"Error: Failed to initialize SQL Gemini model: {model_init_error}"
            logger.error(f"âŒ {error_msg}", exc_info=True)
            llm_response_data = {"type": "text", "data": error_msg}
            results.append(llm_response_data)
            insert_message_into_db({"result": llm_response_data, "chat_id": chat_id, "sender": "System"})
            return results

            # --- Loop Through Tasks ---
        for i, task in enumerate(tasks):
            task_description = task.get("description", f"Unnamed Task {i + 1}")
            task_type = task.get("task_type")
            required_data = task.get(
                "required_data_summary", "No data summary provided"
            )
            viz_type = task.get("visualization_type")

            logger.info(
                f"\n  Task {i + 1}/{len(tasks)}: '{task_description}' ({task_type})"
            )

            if not task_type or not required_data:
                logger.warning(f"  [!] Skipping task - missing type or data summary")
                response = {
                    "type": "error",
                    "data": f"Skipping sub-task '{task_description}' due to incomplete definition from AI.",
                }
                results.append(
                    response
                )
                insert_message_into_db({
                    "chat_id": chat_id,
                    "sender": "System",
                    "result": json.dumps(response)
                })
                continue

            try:
                # --- Step 3a: Generate SQL Query using AI ---
                logger.info(f"    Generating SQL for JSONB...")
                if sql_gemini is None:  # Should not happen if initialization succeeded
                    logger.error("   [âœ—] SQL Gemini model not initialized!")
                    response = {
                        "type": "error",
                        "data": f"Error processing task '{task_description}': SQL generation component not ready.",
                    }
                    results.append(
                        response
                    )
                    insert_message_into_db({
                        "chat_id": chat_id,
                        "sender": "System",
                        "result": json.dumps(response)
                    })
                    continue

                # --- Construct the specific prompt for this task ---
                sql_prompt = (
                    f"Task Description: {task_description}\n"
                    f"Required Data Summary: {required_data}\n"
                    f"Company ID for Query: {company_id}\n"  # Inject the actual company_id
                    f"Generate the PostgreSQL query using ONLY the provided schema"
                    f"querying rules, including the :company_id parameter and correct field access in "
                    f"SELECT/GROUP "
                    f"BY/ORDER BY."
                )  # Added reminder

                sql_chat = sql_gemini.start_chat(history=formated_message_history)
                sql_response = sql_chat.send_message(sql_prompt)
                sql_query_text = clean_response_text(sql_response.text)
                logger.info(
                    f"    Generated SQL:\n{sql_query_text}"
                )  # Log the generated SQL

                # --- Basic SQL Validation ---
                stripped_sql = sql_query_text.lower().strip()
                if not sql_query_text or not (
                        stripped_sql.startswith("select") or stripped_sql.startswith("with")
                ):
                    logger.warning(
                        f"    [âœ—] Invalid or empty SQL query generated by AI (must start with SELECT or WITH): '"
                        f"{sql_query_text}'"
                    )
                    response = {
                        "type": "error",
                        "data": f"Could not generate a valid SQL query (must start with SELECT or WITH) for task: "
                                f"'{task_description}'. AI Output: '{sql_query_text}'",
                    }
                    results.append(
                        response
                    )
                    insert_message_into_db({
                        "chat_id": chat_id,
                        "sender": "System",
                        "result": json.dumps(response)
                    })
                    continue

                logger.info(f"    [âœ“] SQL query generated and basic validation passed.")

                # --- Step 3b: Fetch Data from Database ---
                logger.info(f"    Fetching data using JSONB query...")
                # --- Use the new function with parameter binding ---
                data = sql_query_with_params(
                    sql_query_text,
                    params={"company_id": company_id},
                )

                if not data:
                    # It's possible the query is correct but returns no data matching criteria
                    logger.info(
                        f"    [!] Query executed successfully but returned no data."
                    )
                    response = {
                        "type": "text",
                        "data": f"For '{task_description}': The query executed successfully but found no matching "
                                f"data based on the criteria.",
                    }
                    results.append(
                        response
                    )
                    insert_message_into_db({
                        "chat_id": chat_id,
                        "sender": "System",
                        "result": json.dumps(response)
                    })
                    continue
                else:
                    logger.info(f"    [âœ“] Fetched {len(data)} records")

                # --- Step 3c: Generate Insight, Visualization, or Report ---
                # (No changes needed in this section, it processes the fetched 'data')

                # (Insight Generation Logic)
                if task_type == "insight":
                    logger.info("    Generating insight...")

                    try:
                        # â”€â”€ Call the shared helper â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
                        insight_text = generate_insight(
                            data=data,
                            task_description=task_description,
                        )
                        # Insight generation returns None if it handled an error internally.
                        if insight_text is None:
                            raise RuntimeError("generate_insight returned no output")

                        logger.debug("Generated Insight: %s", insight_text)
                        logger.info("    [âœ“] Insight generated")

                        response = {"type": "text", "data": insight_text}

                    except Exception as err:
                        logger.error("   [âœ—] Insight generation failed: %s", err, exc_info=True)
                        response = {
                            "type": "error",
                            "data": (
                                f"Error processing task '{task_description}': "
                                "Insight generation failed."
                            ),
                        }

                    # â”€â”€ Persist & return the response â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
                    results.append(response)
                    insert_message_into_db({
                        "chat_id": chat_id,
                        "sender": "System",
                        "result": json.dumps(response)
                    })

                # (Visualization Generation Logic)
                elif task_type == "visualization":
                    viz_type_str = viz_type if viz_type else "chart"
                    logger.info(f"    Generating {viz_type_str} visualization...")

                    if not viz_type or viz_type not in ["bar", "line"]:
                        logger.warning(
                            f"    [!] Invalid or missing visualization type '{viz_type}' specified for task."
                        )
                        response = {
                            "type": "error",
                            "data": f"Skipping visualization for '{task_description}': Invalid or missing chart"
                                    f" type ('{viz_type}'). Requires 'bar' or 'line'.",
                        }
                        results.append(
                            response
                        )
                        insert_message_into_db({
                            "chat_id": chat_id,
                            "sender": "System",
                            "result": json.dumps(response)
                        })
                        continue

                    insight_text = None
                    try:
                        insight_text = generate_insight(
                            data=data,
                            task_description=task_description,
                        )

                        if insight_text is not None:
                            # print("Generated Insight:", insight_text)
                            pass
                        else:
                            print("Insight generation failed. No output returned.")

                    except Exception as e:
                        logger.exception(f"Unexpected error while generating insight: {e}")
                        raise

                    if plotly_gemini is None:
                        plotly_instruction = f""" You are a data visualization expert using Plotly.js. Given a 
                                                dataset (as a JSON list of objects), a description of the desired visualization, 
                                                and the required chart type (bar or line), generate the Plotly JSON configuration (
                                                specifically the 'data' and 'layout' objects).

                                                Rules:
                                                - Create a meaningful title for the chart based on the description. Use the exact column 
                                                names (keys) from the dataset for 'x' and 'y' keys in the data trace(s).
                                                - Ensure the generated JSON is syntactically correct and contains ONLY the 'data' (list) and 
                                                'layout' (object) keys at the top level.
                                                - Map the data fields appropriately to x and y axes based on the description and chart type (
                                                'bar' or 'line'). Infer appropriate axes labels from the data keys if not obvious.
                                                - Generate ALL necessary fields for a basic, valid Plotly chart (e.g., 'type', 'x', 
                                                'y' in trace; 'title' in layout). Add axis titles ('xaxis': {{"title": "X Label"}}, 
                                                'yaxis': {{"title": "Y Label"}}).
                                                - If multiple traces are needed (e.g., comparing two values per category), generate a list of 
                                                trace objects within the 'data' list.
                                                - ONLY output the JSON object starting with `{{` and ending with `}}`.
                                                - Do not include any explanations, comments, code blocks (like ```json), or other text.

                                                Example Output Format:
                                                {{
                                                  "data": [
                                                    {{
                                                      "x": [/* array of x-values */],
                                                      "y": [/* array of y-values */],
                                                      "type": "{viz_type}",
                                                      "name": "Optional Trace Name"
                                                    }}
                                                   ],
                                                  "layout": {{
                                                    "title": "Chart Title Based on Description",
                                                    "xaxis": {{"title": "X-Axis Label"}},
                                                    "yaxis": {{"title": "Y-Axis Label"}}
                                                  }}
                                                }}
                                                """
                        try:
                            plotly_gemini = initialize_gemini_model(
                                system_instruction=plotly_instruction
                            )
                            logger.debug("Plotly Gemini model initialized.")
                        except Exception as model_init_error:
                            logger.error(
                                f"   [âœ—] Failed to initialize Plotly Gemini model: {model_init_error}",
                                exc_info=True,
                            )
                            response = {
                                "type": "error",
                                "data": f"Error processing task '{task_description}': Visualization generation "
                                        f"component failed to initialize.",
                            }
                            results.append(
                                response
                            )
                            insert_message_into_db({
                                "chat_id": chat_id,
                                "sender": "System",
                                "result": json.dumps(response)
                            })
                            continue

                    data_keys = (
                        list(data[0].keys()) if data else []
                    )  # Get keys from first record
                    plotly_prompt = f"""
                                            Dataset (JSON format, keys available: {data_keys}):
                                            {json.dumps(data, indent=2, default=str)}

                                            Visualization Description:
                                            "{task_description}"

                                            Required Chart Type:
                                            "{viz_type}"

                                            Generate the Plotly JSON configuration ('data' and 'layout' objects only):
                                            """
                    plotly_chat = plotly_gemini.start_chat()
                    plotly_response = plotly_chat.send_message(plotly_prompt)
                    plotly_json_text = clean_response_text(plotly_response.text)
                    logger.debug(f"Raw Plotly JSON response: {plotly_response.text}")
                    logger.debug(f"Cleaned Plotly JSON response: {plotly_json_text}")

                    try:
                        # More robust check for valid JSON object string
                        if not (
                                plotly_json_text.startswith("{")
                                and plotly_json_text.endswith("}")
                        ):
                            # Try removing potential leading/trailing garbage if simple cleaning failed
                            match = re.search(r"\{.*}", plotly_json_text, re.DOTALL)
                            if match:
                                plotly_json_text = match.group(0)
                            else:
                                raise ValueError(
                                    "Plotly response is not a valid JSON object string."
                                )

                        plotly_json = json.loads(plotly_json_text)
                        # Basic validation
                        if (
                                not isinstance(plotly_json, dict)
                                or "data" not in plotly_json
                                or "layout" not in plotly_json
                        ):
                            raise ValueError(
                                "Plotly JSON missing 'data' or 'layout' key at the top level, or is not an object."
                            )
                        if not isinstance(plotly_json["data"], list):
                            raise ValueError("Plotly 'data' key must be a list.")
                        if not isinstance(plotly_json["layout"], dict):
                            raise ValueError("Plotly 'layout' key must be an object.")
                        # Optional: Deeper validation of trace structure if needed

                        logger.info(f"    [âœ“] Visualization ({viz_type}) generated")
                        response = {"type": "graph", "data": plotly_json,"insight": insight_text}
                        results.append(response)
                        insert_message_into_db({
                            "chat_id": chat_id,
                            "sender": "System",
                            "result": json.dumps(response)
                        })
                    except (json.JSONDecodeError, ValueError) as e:
                        logger.error(
                            f"    [âœ—] Failed to parse or validate Plotly JSON: {e}",
                            exc_info=False,
                        )  # Keep exc_info=False
                        logger.error(
                            f"    Problematic Plotly JSON text: {plotly_json_text}"
                        )
                        response = {
                            "type": "error",
                            "data": f"Error generating visualization for '{task_description}': Invalid Plotly "
                                    f"configuration received from AI. Details: {e}",
                        }
                        results.append(
                            response
                        )
                        insert_message_into_db({
                            "chat_id": chat_id,
                            "sender": "System",
                            "result": json.dumps(response)
                        })

                # (Report Generation Logic)
                elif task_type == "report":
                    insight_text = None
                    try:
                        insight_text = generate_insight(
                            data=data,
                            task_description=task_description,
                        )

                        if insight_text is not None:
                            print("Generated Insight:", insight_text)
                        else:
                            print("Insight generation failed. No output returned.")

                    except Exception as e:
                        logger.exception(f"Unexpected error while generating insight: {e}")
                        raise
                        # Optionally: re-raise or handle as needed

                    logger.info(f"    Generating Excel report...")
                    try:
                        if (
                                not data
                        ):  # Should have been caught earlier, but double check
                            logger.warning(
                                f"    [!] No data to generate Excel report for '{task_description}'."
                            )
                            response = {
                                "type": "text",
                                "data": f"No data available to generate report: '{task_description}'",
                            }
                            results.append(
                                response
                            )
                            insert_message_into_db({
                                "chat_id": chat_id,
                                "sender": "System",
                                "result": json.dumps(response)
                            })
                            continue

                        df = pd.DataFrame(data)

                        excel_buffer = io.BytesIO()
                        # Use a modern engine like openpyxl
                        df.to_excel(
                            excel_buffer,
                            index=False,
                            sheet_name="ReportData",
                            engine="openpyxl",
                        )
                        excel_buffer.seek(0)

                        # Generate AI title for the filename
                        ai_generated_title = task_description  # Fallback
                        if title_gemini is None:
                            title_instruction = """You are an expert at creating concise, descriptive titles for data 
                                                    reports.
                                                                Given a task description and optionally some of the data's column names, 
                                                                generate a short (3-7 words) title suitable for a filename.
                                                                The title should accurately reflect the report's content. Use underscores 
                                                                instead of spaces.
                                                                Output ONLY the title text. No extra formatting, explanations, or quotation 
                                                                marks.
                                                                Example: If task is "Report of sales per region for Q1" -> 
                                                                "Q1_Sales_by_Region_Report"
                                                                Example: If task is "List active users and their last login" -> 
                                                                "Active_Users_Last_Login"
                                                                """
                            try:
                                title_gemini = initialize_gemini_model(
                                    model_name="gemini-1.5-flash",
                                    system_instruction=title_instruction,
                                )
                                logger.debug(
                                    "Title Gemini model initialized for reports."
                                )
                            except Exception as model_init_error:
                                logger.error(
                                    f"   [âœ—] Failed to initialize Title Gemini model: {model_init_error}",
                                    exc_info=True,
                                )
                                # Continue with fallback title

                        if title_gemini:
                            title_prompt_parts = [
                                f'Task Description:\n"{task_description}"\n'
                            ]
                            if not df.empty:
                                title_prompt_parts.append(
                                    f"Data Columns (first few):\n{list(df.columns)[:5]}\n"
                                )
                            title_prompt_parts.append(
                                "Generate a short, filename-friendly title (3-7 words, use underscores):"
                            )
                            title_prompt = "".join(title_prompt_parts)

                            try:
                                title_chat = title_gemini.start_chat()
                                title_response = title_chat.send_message(title_prompt)
                                generated_title_text = clean_response_text(
                                    title_response.text
                                )
                                # Further clean/validate the title
                                generated_title_text = generated_title_text.replace(
                                    " ", "_"
                                )[
                                                       :100
                                                       ]  # Limit length
                                if generated_title_text and re.match(
                                        r"^\w+$", generated_title_text.replace("_", "")
                                ):
                                    ai_generated_title = generated_title_text
                                    logger.info(
                                        f"    AI Generated Title: {ai_generated_title}"
                                    )
                                else:
                                    logger.warning(
                                        f"    AI generated title was invalid or empty ('{generated_title_text}'), "
                                        f"using fallback."
                                    )
                            except Exception as title_gen_error:
                                logger.error(
                                    f"    Error generating title with AI: {title_gen_error}",
                                    exc_info=False,
                                )
                                # Continue with fallback title

                        # Create a safe filename FROM THE AI TITLE (or fallback task_description)
                        # Replace invalid chars, ensure it's not empty, truncate
                        safe_filename_base = re.sub(
                            r"[^\w-]", "_", ai_generated_title
                        ).strip("_")
                        if not safe_filename_base:
                            safe_filename_base = f"report_task_{i + 1}"
                        filename = f"{safe_filename_base[:50]}.xlsx"  # Truncate further for safety
                        wb = openpyxl.load_workbook(excel_buffer)
                        ws = wb.active

                        # Bold and center header row
                        for cell in ws[1]:
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal="center")

                        # Auto-adjust column widths
                        for col in ws.columns:
                            max_length = 0
                            column_letter = get_column_letter(col[0].column)
                            for cell in col:
                                try:
                                    value = (
                                        str(cell.value)
                                        if cell.value is not None
                                        else ""
                                    )
                                    max_length = max(max_length, len(value))
                                except:
                                    pass
                            ws.column_dimensions[column_letter].width = max_length + 2

                        # Optional: Number formatting
                        for row in ws.iter_rows(min_row=2):
                            for cell in row:
                                if isinstance(cell.value, int):
                                    cell.number_format = (
                                        "#,##0"  # For integers (no decimal)
                                    )
                                elif isinstance(cell.value, float):
                                    cell.number_format = (
                                        "#,##0.00"  # For floats (two decimals)
                                    )
                        # ðŸ”¥ Bold the last row (total row)
                        last_row_idx = ws.max_row
                        for cell in ws[last_row_idx]:
                            cell.font = Font(bold=True)

                        # Save formatted workbook back to bytes
                        formatted_buffer = io.BytesIO()
                        wb.save(formatted_buffer)
                        formatted_buffer.seek(0)
                        excel_bytes = formatted_buffer.getvalue()
                        # # Save to local directory
                        # output_path = os.path.join("output", "PM_KPI_Report.xlsx")  # Customize folder/filename
                        # os.makedirs(os.path.dirname(output_path), exist_ok=True)   # Ensure folder exists
                        # with open(output_path, "wb") as f:
                        #     f.write(excel_bytes)
                        excel_base64 = base64.b64encode(excel_bytes).decode("utf-8")

                        logger.info(
                            f"    [âœ“] Excel report '{filename}' prepared (base64 encoded)."
                        )
                        response = {
                            "type": "excel_file",
                            "data": {
                                "filename": filename,
                                "content_base64": excel_base64,
                                "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                "insight": insight_text
                            }
                        }
                        results.append(
                            response
                        )
                        insert_message_into_db({
                            "chat_id": chat_id,
                            "sender": "System",
                            "result": json.dumps(response)
                        })

                    except Exception as report_err:
                        logger.error(
                            f"    [âœ—] Failed to generate Excel report for '{task_description}': {report_err}",
                            exc_info=True,
                        )
                        results.append(
                            {
                                "type": "error",
                                "data": f"Error preparing Excel report for '{task_description}': {report_err}",
                            }
                        )

                else:
                    logger.warning(f"    [!] Unknown task type '{task_type}'")
                    response = {
                        "type": "error",
                        "data": f"Unknown task type '{task_type}' encountered for sub-task '{task_description}'. "
                                f"Cannot process.",
                    }
                    results.append(
                        response
                    )
                    insert_message_into_db({
                        "chat_id": chat_id,
                        "sender": "System",
                        "result": json.dumps(response)
                    })

            except Exception as task_error:
                # Log the specific SQL query that failed if the error is likely SQL related
                if (
                        isinstance(task_error, Exception)
                        and "database" in str(task_error).lower()
                ):
                    logger.error(
                        f"    [âœ—] Database error processing task '{task_description}'. Failed "
                        f"Query:\n{sql_query_text}\nError: {task_error}",
                        exc_info=True,
                    )
                else:
                    logger.error(
                        f"    [âœ—] Error processing task '{task_description}': {task_error}",
                        exc_info=True,
                    )
                response = {
                    "type": "error",
                    "data": f"An error occurred while processing sub-task '{task_description}': {task_error}",
                }
                results.append(
                    response
                )
                insert_message_into_db({
                    "chat_id": chat_id,
                    "sender": "System",
                    "result": json.dumps(response)
                })
            # End of task processing try-except block
            # End of loop through tasks

    except Exception as e:
        logger.error(
            f"âŒ An unexpected error occurred during the main processing pipeline: {e}",
            exc_info=True,
        )
        # Append a generic error message to results if appropriate
        response = {
            "type": "error",
            "data": f"An unexpected error occurred during processing: {e}",
        }
        results.append(
            response
        )
        insert_message_into_db({
            "chat_id": chat_id,
            "sender": "System",
            "result": json.dumps(response)
        })
        # End of main try-except block

    logger.info("\nðŸ PIPELINE EXECUTION COMPLETE")
    if results:
        logger.info(f"Returning {len(results)} result items.")
        # Log types of results generated
        result_types = [r.get("type", "unknown") for r in results]
        logger.info(f"Result types: {', '.join(result_types)}")
    else:
        logger.info("No results were generated.")
    return results